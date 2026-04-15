import threading
import time
#
# Core Keras libraries
#
import shutil
import platform
import dill
#if(platform.system().lower().find('linux')>-1):
#    import autosklearn
import sklearn
# from pyod.models.deep_svdd import DeepSVDD as SVDD
# from pyod.models.ocsvm import OCSVM
import tensorflow
# if(tensorflow.__version__.find('2.8')>-1):
#     import autokeras
import keras
if(keras.__version__>='2.8.0'):
    from keras.utils.vis_utils import plot_model
    from tensorflow.keras.optimizers import Adam as Adam
    from tensorflow.keras.optimizers import SGD as SGD
else:
    from keras.utils import plot_model
    from keras.optimizers import Adam as Adam
    from keras.optimizers import SGD as SGD
from keras.models import Sequential
import keras.layers as keras_layers
from keras.layers import Concatenate, concatenate
from keras import backend as bcd #转换为张量
from keras import callbacks
from keras import models
from keras.callbacks import EarlyStopping 
from keras.layers import Input
from keras.layers import Activation
from keras.models import Model
import tensorflow as tf
import keras.optimizers as omz
from keras.utils.generic_utils import get_custom_objects
from keras import backend as K
import focal_loss
from focal_loss import BinaryFocalLoss
from keras.layers import ReLU
from keras.activations import relu, elu
#
# For data conditioning
#
import pickle
from keras.models import model_from_json
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import confusion_matrix
from sklearn.preprocessing import LabelEncoder as LBC
from sklearn.preprocessing import OneHotEncoder as OHC
from sklearn.metrics import cohen_kappa_score as cohkpa
from sklearn.metrics import mean_squared_log_error
from scipy.ndimage import gaussian_filter1d
from scipy.signal import medfilt
from numpy.random import seed
from numpy.random import choice
import sys
import joblib
# 
# Other essential libraries
#
from copy import deepcopy as dcp
import numpy as np
from numpy import argmax
from numpy import argmin
import pandas as pd
from sklearn.preprocessing import StandardScaler as Stdscr
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn import metrics as skm
from datetime import datetime as dt
import os
#import Config_autosk

from package import visualization as vs
from package import visualization2 as vs2
from package import dataframeprocedure as DFP
from package import LOGger
from package.LOGger import CreateContainer, CreateFile, addlog, copyfile
from package.LOGger import stamp_process, exception_process, for_file_process, set_dir, abspath, mylist, strictly_list, get_classbasename
import time
from package import modeling_core as MDC
vs3=MDC.vs3
from package import algorithms as ALG
import abc


from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#%%
m_debug = LOGger.myDebuger(stamps=[*os.path.basename(__file__).split('.')[:-1]])
parseKL = lambda nn_type, **kwags: getattr(keras.layers, nn_type, None)

def method_activation(stg):
    try:
        method = eval(stg)
        return method
    except:
        print('method invalid:%s!!!!'%stg)
        return None
######################################################################################################################################
# abc_EIMS
######################################################################################################################################
def checkConfigAvailable(config_file):
    if(not LOGger.isinstance_not_empty(config_file, str)):
        print("config_file err: %s"%config_file)
        return False
    if(not os.path.exists(config_file)):
        print("config_file doesn't exists: %s"%config_file)
        return False
    return True

def configFromFile(config_file):
    if(not checkConfigAvailable(config_file)):
        return {}
    config = LOGger.load_json(config_file)
    return config

def loadConfigFromFileStandard(mdl, config_file, eva_bck='$', **kwags):
    config = configFromFile(config_file)
    if(not mdl.loadConfig(config, eva_bck=eva_bck)):
        return False
    return True
    
def loadConfigStandard(mdl, config, eva_bck='$', **kwags):
    mdl.config = {}
    for k,v in config.items():
         v = dcp(LOGger.mystr(v).config_evaluation(eva_bck=eva_bck, **config) if(isinstance(v, str)) else v)
         mdl.config[k] = v
    return True

def setProperty(mdl, topic, default=None):
    if(not hasattr(mdl, '%sProperty'%topic.lower())):
        mdl.addlog("no 'propertyTopic!!!!!!!!", colora=LOGger.FAIL)
        return False
    # LOGger.addDebug(str(mdl.config))
    properties = getattr(mdl, '%sProperty'%topic.lower())
    for k in properties:
        pDefault = dcp(default)
        if(k in properties):   pDefault = dcp(properties[k])
        p = mdl.config.get(k, pDefault)
        if(LOGger.isinstance_not_empty(p, str)):
            p = LOGger.mystr(p).config_evaluation(**mdl.config)
        setattr(mdl, k, p)
        mdl.addlog("set %s:%s"%(k, DFP.parse(p)), stamps=[topic])
    return True

def setInitProperty(mdl, **kwags):
    """
    setProperty, setOuputProperty, mdc, inheritPropToCore...

    Parameters
    ----------
    mdl : TYPE
        DESCRIPTION.

    Returns
    -------
    bool
        DESCRIPTION.

    """
    if(LOGger.isinstance_not_empty(mdl.config, dict)):  
        if(not mdl.loadConfig(config=mdl.config)):
            mdl.addlog('loadConfig failed!!!', colora=LOGger.FAIL)
            return False
    elif(LOGger.isinstance_not_empty(mdl.config_file, str)):  
        if(not mdl.loadConfigFromFile(file=mdl.config_file)):
            mdl.addlog('loadConfigFromFile failed!!!', colora=LOGger.FAIL)
            return False
    setProperty(mdl, 'supre')
    setProperty(mdl, 'core')
    setProperty(mdl, 'data')
    if(not setInputProperty(mdl, **mdl.dataProperty)):
        mdl.addlog('setInputProperty failed!!!', colora=LOGger.FAIL)
        return False
    if(not setOuputProperty(mdl)):
        mdl.addlog('setOuputProperty failed!!!', colora=LOGger.FAIL)
        return False
    if(not inheritPropToCore(mdl)):
        return False
    if(not mdl.config.get('pre_model_module')):
        if(not preprocessingScenario(mdl)):
            mdl.mdc.addlog('preprocessingScenario failed!!!', stamps=mdl.mdc.get_stamps(), colora=LOGger.FAIL)
            return False
    if(mdl.need_training):
        if(not saveModules(mdl)):
            mdl.mdc.addlog('saveModules failed!!!', stamps=mdl.mdc.get_stamps(), colora=LOGger.FAIL)
            return False
    return True

def setOuputProperty(mdl):
    """
    set exp_fd, logfile, addlog, ...

    Parameters
    ----------
    mdl : TYPE
        DESCRIPTION.

    Returns
    -------
    bool
        DESCRIPTION.

    """
    datetimestg = dt.now().strftime('%Y%m%d')
    if(not LOGger.isinstance_not_empty(getattr(mdl,'exp_fd',''), str)):
        mdl.exp_fd = DFP.pathrpt(os.path.join('aftdata', *mdl.theme_layers, datetimestg, 'round'))
    if(not constructOutputDiretory(mdl)):
        return False
    mdl.logfile = os.path.join(mdl.exp_fd, 'log.txt')
    mdl.addlog = LOGger.addloger(logfile=mdl.logfile)
    return True

def setInputProperty(mdl, stamps=None, **kwags):
    """
    set source_data, ...

    Parameters
    ----------
    mdl : TYPE
        DESCRIPTION.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    stamps = stamps if(isinstance(stamps, list)) else []
    mdl.source_data = eval(mdl.importMethod)(mdl.source_data_file, **kwags)
    dataCount = mdl.source_data.shape[0]
    mdl.addlog('dataCount:', dataCount, stamps=stamps)
    dataCountLbd = getattr(mdl, 'dataCountLbd', 0)
    return mdl.source_data.shape[0]>=dataCountLbd
    
def constructOutputDiretory(mdl):
    if(getattr(mdl, 'rewrite', False)):
        shutil.rmtree(mdl.exp_fd) if(os.path.isdir(mdl.exp_fd) and mdl.exp_fd!='.') else None
    if(not os.path.isdir(mdl.exp_fd)):
        LOGger.CreateContainer(mdl.exp_fd)
    return True

def inheritPropToCore(mdl):
    if(not mdl.need_training and not mdl.config.get('pre_model_module')):
        mdl.addlog('No training and no pre_model_module!!!!!', colora=LOGger.FAIL)
        return False
    if('pre_model_module' in mdl.config):    
        mdl.addlog('pre_model module:', mdl.config['pre_model_module'], colora=LOGger.OKCYAN)
        mdl.mdc = getattr(MDC, mdl.coreClassName, 'EIMS_core')(
            exp_fd=mdl.exp_fd, version=mdl.version, config_file=os.path.join(mdl.config.get('pre_model_module','.'), 'config.json'),
            stamps=mdl.stamps)
        if(not mdl.mdc.set_config()):
            return False
        if(mdl.need_training):
            mdl.config['version'] = mdl.config['version'] + '-1'
            mdl.addlog('version updated:', mdl.config['version'], colora=LOGger.OKCYAN)
        
        config = dcp(mdl.config)
        for k,v in mdl.coreProperty.items():
            config[k] = config.get(k, v)
        mdl.mdc.update_config(config=config)
        # sys.exit(1)
    else:
        mdl.mdc = getattr(MDC, mdl.coreClassName, 'EIMS_core')(exp_fd=mdl.exp_fd, version=mdl.version)
        if(not  mdl.mdc.set_config(config=mdl.config)):
            return False
    return True

def saveModules(mdl, common_exp_fd=None, **kwags):
    """
    |||   ...mdc.save_modules(exp_fd=common_exp_fd)...   |||
    mdc.save_models -> save_mdcHeaderPreprocessors -> save_config
    If exp_fd==None, `mdc.save_models`,`save_mdcHeaderPreprocessors` will be save following mdc.get_module_exp_fd(), while save_config will be saved follow mdc.exp_fd

    Parameters
    ----------
    mdl : TYPE
        DESCRIPTION.
    common_exp_fd : TYPE
        DESCRIPTION. The default is None.

    Returns
    -------
    bool
        DESCRIPTION.

    """
    mdc = mdl.mdc
    if(not mdc.save_modules(exp_fd=common_exp_fd)):
        mdc.addlog('mdc save_modules failed!!!', stamps=mdc.get_stamps(), colora=LOGger.FAIL)
        return False
    return True

def preprocessingScenario(mdl):
    mdc = mdl.mdc
    data = mdl.source_data
    for header_zones_name in mdc.IOheader_zones_names:
        header_zones = getattr(mdc, header_zones_name)
        mdc.addlog('headerZones count:', len(header_zones), stamps=[header_zones_name])
        for headerZoneName, headerZone in header_zones.items():
            overHeader = [x for x in headerZone if x not in data.columns]
            if(len(overHeader)>0):
                mdc.addlog('source_data_file:', str(mdl.source_data_file), stamps=mdc.get_stamps()+[headerZoneName])
                mdc.addlog('headerZone:', str(headerZone), stamps=mdc.get_stamps()+[headerZoneName])
                mdc.addlog('data.columns:', str(list(data.columns)), stamps=mdc.get_stamps()+[headerZoneName])
                mdc.addlog('overHeader', overHeader, colora=LOGger.FAIL)
                mdc.addlog('headerZone not match data.columns', stamps=mdc.get_stamps()+[headerZoneName], colora=LOGger.FAIL)
                return False
            if(not headerZone.fit(referenceData=data[headerZone])):
                mdc.addlog('headerZone fit failed!!!', stamps=mdc.get_stamps()+[headerZoneName], colora=LOGger.FAIL)
                return False
            mdc.addlog('headerZone fit success!!!', stamps=[header_zones_name, headerZoneName], colora=LOGger.OKGREEN)
    return True

######################################################################################################################################
# 
######################################################################################################################################
def get_hyperlink(mdl):
    if(getattr(mdl, 'exp_fd', '')):
        exp_fd = mdl.exp_fd
        exp_fd_rel = os.path.relpath(exp_fd, start=mdl.project_exp_fd)
        round_title = ('%s'%exp_fd_rel[exp_fd_rel.find('round')+6:] if(
                exp_fd_rel.find('round')+6<len(exp_fd_rel)) else '0') if(
                exp_fd_rel.find('round')>-1) else os.path.basename(mdl.mdc.exp_fd)
        return '=HYPERLINK("%s", "%s")'%(exp_fd_rel, round_title)

def produce_stratify(mdl, data=None, **kwags):
    mdc = mdl.mdc
    tn_ratio = getattr(mdc, 'tn_ratio', getattr(mdl, 'tn_ratio', 0.85))
    data = data if(isinstance(data, pd.core.frame.DataFrame)) else mdl.source_data
    stratifyLabels = mdl.stratifyLabels
    print('tn_ratio:', tn_ratio)
    print('stratifyLabels\n', stratifyLabels)
    stratify = None
    parse_method = lambda x:DFP.parse(x,1)
    if(LOGger.isinstance_not_empty(stratifyLabels, list)):
        stratify = [DFP.parse(v) if(np.array(v).shape==()) else ' | '.join(list(map(parse_method, v)))
                    for v in data[stratifyLabels].values]
        stratify = DFP.adjust_stratify(
                stratify, class_max=(1-tn_ratio)*len(data.index))
    return stratify

def train_test_logging(mdl, **kwags):
    # mdc = mdl.mdc
    shape_stamps = {}
    shape_stamps.update({'X_train':'%s'%str(mdl.X_train.shape)}) if(not mdl.X_train.empty) else None
    shape_stamps.update({'y_train':'%s'%str(getattr(mdl,'y_train',None).shape)}) if(not getattr(mdl,'y_train',None).empty) else None
    shape_stamps.update({'X_test':'%s'%str(mdl.X_test.shape)}) if(not mdl.X_test.empty) else None
    shape_stamps.update({'y_test':'%s'%str(getattr(mdl,'y_test',None).shape)}) if(not getattr(mdl,'y_test',None).empty) else None
    log = stamp_process('shapes:', shape_stamps)
    mdl.addlog(log, **kwags)

def dataPurposeClassificationStandard(mdl, df, **kwags):
    mdc = mdl.mdc
    tn_ratio = getattr(mdc, 'tn_ratio', getattr(mdl, 'tn_ratio', 0.85))
    stratify = produce_stratify(mdl, df)
    dfx = df[mdc.xheader]
    dfy = df[mdc.yheader]
    dfa = df[mdc.auxheader]
    mdl.dfx = dfx
    mdl.dfy = dfy
    mdl.dfa = dfa
    X_train, X_test, y_train, y_test, aux_train, aux_test = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if(mdl.need_training):
        try:
            X_train,X_test,y_train,y_test,aux_train,aux_test = train_test_split(
                dfx, dfy, dfa, test_size=1-tn_ratio, random_state=5, stratify=stratify)
        except Exception as e:
            m_debug.update({'dfx': dfx, 'dfy':dfy, 'dfa':dfa, 'stratify':stratify, 'tn_ratio':tn_ratio})
            m_debug.save()
            LOGger.exception_process(e,logfile='')
            mdl.addlog('train_test_split失敗, 請檢查資料是否補缺完全或是有其他疏漏....!!!!', colora=LOGger.FAIL)
            return False
        mdl.X_train = X_train
        discrete_header = mylist([v for v in mdc.yheader_zones.values() if getattr(v, 'data_prop', None) in ['discrete','categorical','binary']]).get_all()
        mdc.addlog(str(discrete_header))
        y_train_discrete = y_train[discrete_header]
        mdl.addlog('y_train classes:\n%s...'%str(DFP.uniqueByIndex(
            np.array(y_train_discrete), axis=(None if(mylist(y_train_discrete.shape).get(1,0)<=1) else 1), return_index=True))[:200]) if(
            not y_train_discrete.empty) else None
        mdl.y_train = y_train
        mdl.aux_train = aux_train
    else:
        X_test = dfx.copy()
        y_test = dfy.copy()
        aux_test = dfa.copy()
    mdl.X_test = X_test
    mdl.y_test = y_test
    mdl.aux_test = aux_test
    train_test_logging(mdl, **kwags)
    return True

def dataPurposeClassificationOnlyCause(mdl, df, **kwags):
    mdc = mdl.mdc
    tn_ratio = getattr(mdc, 'tn_ratio', getattr(mdl, 'tn_ratio', 0.85))
    stratify = produce_stratify(mdl, df)
    dfx = df[mdc.xheader]
    dfa = df[mdc.auxheader]
    mdl.dfx = dfx
    mdl.dfy = dfx
    mdl.dfa = dfa
    X_train, X_test, aux_train, aux_test = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if(mdl.need_training):
        try:
            X_train,X_test,aux_train,aux_test = train_test_split(
                dfx, dfa, test_size=1-tn_ratio, random_state=5, stratify=stratify)
        except Exception as e:
            m_debug.update({'dfx': dfx, 'dfa':dfa, 'stratify':stratify, 'tn_ratio':tn_ratio})
            m_debug.save()
            LOGger.exception_process(e,logfile='')
            mdl.addlog('train_test_split失敗, 請檢查資料是否補缺完全或是有其他疏漏....!!!!', colora=LOGger.FAIL)
            return False
        mdl.X_train = dcp(X_train)
        mdl.y_train = dcp(X_train)
        discrete_header = mylist([v for v in mdc.xheader_zones.values() if 
                                  getattr(v, 'data_prop', None) in ['discrete','categorical','binary']]).get_all()
        mdc.addlog(str(discrete_header))
        X_train_discrete = X_train[discrete_header]
        mdl.addlog('classes:\n%s...'%str(DFP.uniqueByIndex(
            np.array(X_train_discrete), axis=(None if(mylist(X_train_discrete.shape).get(1,0)<=1) else 1), return_index=True))[:200]) if(
            not X_train_discrete.empty) else None
        mdl.aux_train = aux_train
    else:
        X_test = dfx.copy()
        aux_test = dfa.copy()
    mdl.X_test = dcp(X_test)
    mdl.y_test = dcp(X_test)
    mdl.aux_test = dcp(aux_test)
    train_test_logging(mdl, **kwags)
    return True

def compilingStandard(mdl, **kwags):
    mdc = mdl.mdc
    return mdc.compiling(need_training=mdl.need_training, **kwags)

def transformArrays2Tensors(mdl, stamp, headerZoneItems, dataCountLbd=0, default_data_value=0):
    stamp = stamp if(isinstance(stamp, str)) else 'main'
    mdc = mdl.mdc
    data = getattr(mdl, stamp)
    if(not MDC.transformArrays2Tensors(mdc, data, headerZoneItems, stamp=stamp, dataCountLbd=dataCountLbd, default_data_value=default_data_value)):
        return False
    mdl.addlog('\n', type(data), str(data)[:200], stamps=[stamp, 'original'])
    dataTensorType = getattr(mdc, LOGger.stamp_process('',[stamp, 'tensor'],'','','','_'))
    mdl.addlog('\n', type(dataTensorType), str(dataTensorType)[:200], stamps=[stamp, 'tensorized'])
    setattr(mdl, LOGger.stamp_process('',[stamp, 'tensor'],'','','','_'), getattr(mdc, LOGger.stamp_process('',[stamp, 'tensor'],'','','','_')))
    return True

def modelTraining(mdl, is_self_supervise=False, stamps=None, **kwags):
    is_valid_data = mdl.is_valid_data
    stamps = stamps if(isinstance(stamps, list)) else []
    mdc = mdl.mdc
    if(not transformArrays2Tensors(mdl, 'X_train', tuple(mdc.xheader_zones.items()), **kwags)):
        return False
    if(is_valid_data):
        if(not transformArrays2Tensors(mdl, 'X_test', tuple(mdc.xheader_zones.items()), **kwags)):
            return False
    if(not transformArrays2Tensors(mdl, 'y_train', tuple(mdc.yheader_zones.items() if(
        not isinstance(mdc, MDC.EIMS_AUTOENCODER_core)) else mdc.xheader_zones.items())+
                                tuple(getattr(getattr(mdc, 'unfam_header_zones', None), 'items', lambda :())()), **kwags)):
        return False
    if(is_valid_data):
        if(not transformArrays2Tensors(mdl, 'y_test', tuple(mdc.yheader_zones.items() if(
            not isinstance(mdc, MDC.EIMS_AUTOENCODER_core)) else mdc.xheader_zones.items())+
                                    tuple(getattr(getattr(mdc, 'unfam_header_zones', None), 'items', lambda :())()), **kwags)):
            return False
    if(not mdc.fit(mdl.X_train_tensor, mdl.y_train_tensor, validation_data=((mdl.X_test_tensor, mdl.y_test_tensor) if(is_valid_data) else ()), **kwags)):
        mdc.addlog('mdc fit failed!!!', colora=LOGger.FAIL)
        return False
    log = LOGger.stamp_process('',['訓練結束', mdc.getStringModelInOut()], '','','','_', colora=LOGger.OKCYAN)
    mdc.addlog(log, stamps=mdc.get_stamps())
    mdc.save_models(theme=mdl.createModelFileBasename().split('.')[0])
    if(not mdc.unfam_header_zones):
        return True
    unfamScores = mdc.predict_score(mdl.X_train, zone_index=getattr(mdc, 'unfam_header_zone_index', -1))
    mdc.threshold_unfam = MDC.determined_threshold(mdc, unfamScores, **kwags)
    mdc.addlog('widely used threshold_unfam:%s'%mdc.threshold_unfam, stamps=mdc.get_stamps()+['unfam_prop'], colora=LOGger.OKGREEN)
    return True

def fitStandard(mdl):
    if(not MDC.compilingScenario(mdl.mdc, need_training=True)):
        return False
    if(not modelTraining(mdl)):
        return False
    return True

def createModelFileBasename(mdl, **kwags):
    mdc = mdl.mdc
    return '%s.%s'%(LOGger.stamp_process('',[mdl.theme]+mdc.get_stamps(),'','','','_',for_file=True),  mdc.modelExpfileType)

# def configuringModelNames(mdl, **kwags):
#     for 

def unfamEvaluationScenario(mdl, X_data, ret=None, handler=None, qualifiedMask=None, stamps=None, **kwags):
    mdc = mdl.mdc
    if(not MDC.unfamEvaluationScenario(mdc, X_data, handler=handler, qualifiedMask=qualifiedMask, stamps=stamps, **kwags)):
        return False
    return True
    

def evaluationSimple(mdl, XDataStamp, yDataStamp, **kwags):
    mdc = mdl.mdc
    X_data = getattr(mdc, XDataStamp)
    y_data = getattr(mdc, yDataStamp)
    if(not mdc.evaluation(X_data, y_data, **kwags)):
        return False

def evaluationEquipBasicData(df):
    """
    equiping cause data, aux data...

    Parameters
    ----------
    df : TYPE
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    """
    return MDC.evaluationEquipBasicData(df)

def evaluationStandard(mdl, X_data, y_data, aux_data=None, handler=None, stamps=None, 
                       figUnfamScores=None, determine_threshold=False, axAddingLayout=(2,1), **kwags):
    stamps = stamps if(isinstance(stamps, list)) else []
    stamp = LOGger.stamp_process('',stamps,'','','','_')
    mdc = mdl.mdc
    handler.eval_data = pd.DataFrame()
    retTemp = {}
    kwags['ret'] = retTemp
    handler.success = mdc.evaluation(X_data, y_data, stamps=stamps, handler=handler, determine_threshold=determine_threshold,
                                     figUnfamScores=figUnfamScores, axAddingLayout=axAddingLayout, **kwags)
    mdl.export[stamp] = {}
    header_zones = mdl.mdc.yheader_zones if(not isinstance(mdl.mdc, MDC.EIMS_AUTOENCODER_core)) else mdl.mdc.xheader_zones
    for k,v in header_zones.items():
        mdl.export[stamp][k] = dcp(v.data_prop.export.get('main', v.data_prop.export))
    handler.eval_data = dcp(handler.eval_data.join(evaluationEquipBasicData(X_data), sort=False))
    if(not getattr(aux_data, 'empty', True)):   
        handler.eval_data = dcp(handler.eval_data.join(evaluationEquipBasicData(aux_data), sort=False))
    if(mdc.unfam_header_zones):
        qualifiedMask = retTemp.get('mask', np.full(X_data.shape[0], True))
        ax = figUnfamScores.add_subplot(*axAddingLayout,len(vs3.get_frames(figUnfamScores))+1) if(
            isinstance(figUnfamScores, vs3.plt.Figure)) else None
        if(not unfamEvaluationScenario(mdl, X_data, handler=handler, qualifiedMask=qualifiedMask, stamps=stamps,
                                       determine_threshold=determine_threshold, ax=ax)):
            return False
    return True

def dyeingExcel(dataCol, mask, ws):
    # 應用顏色到 A 欄位
    for row_idx, (value, color) in enumerate(zip(dataCol, mask), start=2):  # 從 Excel 第 2 列開始
        cell = ws[f'A{row_idx}'] #row index要想清楚
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return True

def interpolate_color(val, a, b, start_color=(255, 255, 255), end_color=(255, 0, 0)):
    """計算數值 val 在 a~b 區間內的線性顏色變化"""
    if val <= a:
        return f"{start_color[0]:02X}{start_color[1]:02X}{start_color[2]:02X}"  # 起始色
    elif val >= b:
        return f"{end_color[0]:02X}{end_color[1]:02X}{end_color[2]:02X}"  # 終點色
    else:
        ratio = (val - a) / (b - a)
        r = int(start_color[0] + (end_color[0] - start_color[0]) * ratio)
        g = int(start_color[1] + (end_color[1] - start_color[1]) * ratio)
        b = int(start_color[2] + (end_color[2] - start_color[2]) * ratio)
        return f"{r:02X}{g:02X}{b:02X}"

def dyeingExcelFromFile(source_file, mask=None, maskColIndex=0, maskColName=None):
    # 讀取 Excel 檔案並修改背景顏色
    wb = load_workbook(source_file)
    ws = wb.active
    max_row = ws.used_range.rows.count  # 有內容的最大行
    max_col = ws.used_range.columns.count  # 有內容的最大列
    
    table = pd.DataFrame(ws.range((2,max_col),(2,max_row)).value, columns=ws.range(1,max_col), index=(2,max_row))
    if(mask is None):   
        if(isinstance(maskColName, str)):
            maskCol = table[maskColName]
        else:
            maskCol = table[maskColIndex]
        mask = maskCol.apply(lambda x: interpolate_color(x, a, b, start_color, end_color)).copy()
    if(not dyeingExcel(maskCol, mask, ws)):
        return False
    
    # 儲存修改後的 Excel 檔案
    wb.save(source_file)
    return True

def evaluationScenario(mdl, **kwags):
    """
    mdl evaluate train data if mdl need training; evaluate testing data; evaluate unfam data if neccesary....

    Parameters
    ----------
    mdl : TYPE
        DESCRIPTION.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    """
    figUnfamScores = vs3.plt.Figure(figsize=mdl.figUnfamScoresFigsize) if(hasattr(mdl, 'figUnfamScoresFigsize')) else None
    figUnfamScoresLayout = mdl.figUnfamScoresLayout if(hasattr(mdl, 'figUnfamScoresLayout')) else (2,1)
    
    mdc = mdl.mdc
    mdl.handlerTestData = LOGger.mystr()
    mdl.handlerTrainData = LOGger.mystr()
    if(mdl.need_training):
        mdl.evaluation(mdl.X_train, mdl.y_train, aux_data=mdl.aux_train, handler=mdl.handlerTrainData, stamps=['train'],  
                       figUnfamScores=figUnfamScores, determine_threshold=True, axAddingLayout=mdl.figUnfamScoresLayout, **kwags)
    else:
        mdl.handlerTrainData.success = True
    mdl.evaluation(mdl.X_test, mdl.y_test, aux_data=mdl.aux_test, handler=mdl.handlerTestData, stamps=['test'],
                   figUnfamScores=figUnfamScores, determine_threshold=False, 
                   axAddingLayout=((figUnfamScoresLayout[0] if(mdl.need_training) else 1),figUnfamScoresLayout[1]), **kwags)
    
    file_dtl = os.path.join(mdc.get_detail_exp_fd(),'%s.xlsx'%LOGger.stamp_process(
        '', [mdl.theme]+mdc.get_stamps(for_file=True)+['detailed','evaluation'],'','','','_'))
    CreateContainer(file_dtl)
    wrt_dtl = pd.ExcelWriter(file_dtl, engine='xlsxwriter')
    if(mdl.need_training):  mdl.handlerTrainData.eval_data.to_excel(wrt_dtl, 'train')
    mdl.handlerTestData.eval_data.to_excel(wrt_dtl, 'test')
    CreateFile(file_dtl, lambda f:wrt_dtl.save())
    
    if(mdc.unfam_header_zones):
        fileUnfamScore = os.path.join(mdc.get_graph_exp_fd(), '%s.jpg'%LOGger.stamp_process('',mdc.get_stamps()+['plotUnfamScores'],'','','','_',for_file=True))
        LOGger.CreateFile(fileUnfamScore, lambda f:vs3.end(figUnfamScores, file=f))
    
    MDC.encodingKerasLattenMapping_target_threading(
        mdc, fig=getattr(mdc,'figLattenSpaces',None), 
        figsize=getattr(mdc,'lattenSpacesFigsize',(40,40)), stamps=mdc.get_stamps(), **kwags)
    return mdl.handlerTrainData.success & mdl.handlerTestData.success

def projectRecordOfGeneralHeaderStandard(mdl, **kwags):
    mdc = mdl.mdc
    dic = {}
    dic['type'] = '%s:%s'%(LOGger.type_string(mdl), LOGger.type_string(mdl.mdc))
    dic['hyperlink'] = mdl.get_hyperlink()
    dic['version'] = mdc.version
    dic['test_count'] = int(mdl.y_test.shape[0])
    dic['train_count'] = int(mdl.y_train.shape[0]) if(not mdl.y_train.empty) else 0
    dic['source'] = LOGger.make_hyperlink(os.path.relpath(os.path.dirname(mdl.source_data_file), start=mdl.project_exp_fd), 
                                          title=os.path.basename(mdl.source_data_file))
    xheader, yheader = mdc.xheader, mdc.yheader
    xheader_count = np.array(xheader).shape[0]
    yheader_count = np.array(yheader).shape[0]
    dic['xheader'] = '%s:%s'%(xheader_count, 
       ','.join([DFP.parse(v) for v in xheader[:30]]+(['..'] if(xheader_count>30) else [])))
    dic['yheader'] = '%s:%s'%(yheader_count, 
       ','.join([DFP.parse(v) for v in yheader[:30]]+(['..'] if(yheader_count>30) else [])))
    dic['source_model'] = mdc.get_hyperlink()
    return dic

#TODO:projectEvaluationScenario
def projectRecordOfRegressionHeader(infrm, stamps=None, **kwags):
    stamps = stamps if(isinstance(stamps, list)) else []
    infrm = dcp(infrm.get('main', infrm))
    dic = {}
    dic[stamp_process('',stamps+['RMSE'],'','','','_')] = '%s'%DFP.parse(infrm.get('rmse',-1), digit=4)
    dic[stamp_process('',stamps+['tol'],'','','','_')] = '%s'%DFP.parse(infrm.get('tol',-1), digit=4)
    dic[stamp_process('',stamps+['OKR'],'','','','_')] = '%s'%DFP.parse(infrm.get('OKR',-1), digit=4)
    dic[stamp_process('',stamps+['r2'],'','','','_')] = '%s'%DFP.parse(infrm.get('r2',-1), digit=6)
    dic[stamp_process('',stamps+['adr2'],'','','','_')] = '%s'%DFP.parse(infrm.get('adr2',-1), digit=6)
    return dic

def projectRecordOfRegressionHeaderByDimInZone(infrm, zone, stamps=None, **kwags):
    stamps = stamps if(isinstance(stamps, list)) else []
    zone = LOGger.mylist(list(tuple(zone)) if(DFP.isiterable(infrm)) else [])
    infrm = dcp(infrm.get('main', infrm))
    dic = {}
    dic[stamp_process('',stamps+['RMSE'],'','','','_')] = '%s'%DFP.parse(infrm.get('rmse',-1), digit=4) if(
        len(zone)==1) else stamp_process('', {yn:'%s'%DFP.parse(infrm[
        yn].get('rmse',-1)) for yn in zone}, ':','','',', ')
    dic[stamp_process('',stamps+['tol'],'','','','_')] = '%s'%DFP.parse(infrm.get('tol',-1), digit=4) if(len(zone)==1
                            ) else stamp_process('', {yn:'%s'%DFP.parse(infrm[
                            yn].get('tol',-1)) for yn in zone}, ':','','',', ')
    dic[stamp_process('',stamps+['OKR'],'','','','_')] = '%s'%DFP.parse(infrm.get('OKR',-1), digit=4) if(len(zone)==1
                            ) else stamp_process('', {yn:'%s'%DFP.parse(infrm[
                            yn].get('OKR',-1)) for yn in zone}, ':','','',', ')
    dic[stamp_process('',stamps+['r2'],'','','','_')] = '%s'%DFP.parse(infrm.get('r2',-1), digit=6) if(len(zone)==1
                            ) else stamp_process('', {yn:'%s'%DFP.parse(infrm[
                            yn].get('r2',-1)) for yn in zone}, ':','','',', ')
    dic[stamp_process('',stamps+['adr2'],'','','','_')] = '%s'%DFP.parse(infrm.get('adr2',-1), digit=6) if(
        len(zone)==1) else stamp_process('', {yn:'%s'%DFP.parse(infrm[yn].get('adr2',-1)) for yn in zone}, ':','','',', ')
    return dic

def projectRecordOfClassificationHeader(infrm, stamps=None, f_beta=None, threshold_binary=None, **kwags):
    stamps = stamps if(isinstance(stamps, list)) else []
    # classification_infrm = zone.data_prop.export
    dic = {}
    dic[stamp_process('',stamps+['acc'],'','','','_')] = DFP.parse(infrm.get('acc', ''), 2)
    dic[stamp_process('',stamps+['$\kappa$'],'','','','_')] = DFP.parse(infrm.get('kappa', ''), 2)
    dic[stamp_process('',stamps+['pcs_mdn'],'','','','_')] = DFP.parse(infrm.get('pcs_mdn', ''), 2)
    dic[stamp_process('',stamps+['pcs'],'','','','_')] = DFP.parse(infrm.get('pcs', ''), 2)
    dic[stamp_process('',stamps+['rcl'],'','','','_')] = DFP.parse(infrm.get('rcl', ''), 2)
    dic[stamp_process('',stamps+['f-score'],'','','','_')] = DFP.parse(infrm.get('f-score', ''), 2)
    dic[stamp_process('',stamps+['alpha'],'','','','_')] = DFP.parse(infrm.get('e_type1', ''), 2)
    dic[stamp_process('',stamps+['beta'],'','','','_')] = DFP.parse(infrm.get('e_type2', ''), 2)
    dic[stamp_process('',stamps+['threshold_binary'],'','','','_')] = DFP.parse(infrm.get('threshold_binary', 4))
    dic[stamp_process('',stamps+['f_beta'],'','','','_')] = DFP.parse(infrm.get('f_beta', f_beta), 2)
    return dic

def projectRecordOfParitcularHeader(mdc, propertiesClass, stamps=None, **kwags):
    stamps = stamps if(isinstance(stamps, list)) else []
    dic = {}
    for k in getattr(mdc,'%sAttrs'%propertiesClass,[]):
        dic[stamp_process('',stamps+[k],'','','','_')] = DFP.parse(getattr(mdc, k, ''), 4)
    return dic

def projectRecordStandard(mdl, **kwags):
    mdc = mdl.mdc
    header_zones = mdc.yheader_zones if(not isinstance(mdc, MDC.EIMS_AUTOENCODER_core)) else mdc.xheader_zones
    mdl.config_dict = {}
    mdl.config_dict.update(mdl.projectRecordOfGeneralHeader())
    for k,v in header_zones.items():
        data_prop = getattr(v, 'data_prop', 'continuous')
        infrm = mdl.export['test'][k]
        if(data_prop=='continuous'):
            mdl.config_dict.update(projectRecordOfRegressionHeader(infrm, stamps=['test', k], zone=v))
        elif(data_prop in ['discrete','binary','categorical']):
            mdl.config_dict.update(projectRecordOfClassificationHeader(infrm, stamps=['test', k]))
    if(mdl.config.get('memo','')):
        mdl.config_dict['memo'] = mdl.config['memo']
    if(mdl.need_training):
        mdl.config_dict.update(projectRecordOfParitcularHeader(mdc, 'compile'))
        for k,v in header_zones.items():
            data_prop = getattr(v, 'data_prop', 'continuous')
            infrm = mdl.export['train'][k]
            if(data_prop=='continuous'):
                mdl.config_dict.update(projectRecordOfRegressionHeader(infrm, stamps=['train', k], zone=v))
            elif(data_prop in ['discrete','binary','categorical']):
                mdl.config_dict.update(projectRecordOfClassificationHeader(infrm, stamps=['train', k]))
    return True

def saveProjectRecord(mdl, stamps=None, **kwags):
    mdc = mdl.mdc
    stamps = stamps if(isinstance(stamps, list)) else (mdl.exp_fd if(mdl.exp_fd.find('test')>-1) else mdc.get_stamps(full=True))
    stamps = stamps if(LOGger.isinstance_not_empty(stamps, list)) else ['main']
    # LOGger.addDebug('stamps', stamps)
    exp_fd = mdl.exp_fd
    dic = mdl.config_dict
    prm_fn = getattr(mdl,'prm_fn','%s.xlsx'%LOGger.stamp_process('', ['records',mdl.theme], '','','','_'))
    log = '讀取參數檔案:%s.....'%(prm_fn)
    addlog_ = kwags.get('addlog', LOGger.addloger(logfile=os.path.join(exp_fd, 'log.txt')))
    addlog_(log, stamps=stamps)
    evl = DFP.evaluation()
    try:
        evl.add_export(dic, prm_fn, sheet_name = LOGger.stamp_process('',stamps,'','','','_') if(
            not LOGger.isinstance_not_empty(mdl.config.get('exp_fd'),str)) else os.path.basename(mdl.exp_fd))
    except Exception as e:
        LOGger.exception_process(e, stamps=stamps, logfile='', **kwags)
        m_debug['stamps'] = stamps
        m_debug['dic'] = dic
        m_debug['prm_fn'] = prm_fn
        m_debug['storageboxes'] = evl.storageboxes
        m_debug['evl'] = evl
        m_debug.save()
        return False
    if(not LOGger.save_json(mdl.config, os.path.join(exp_fd, os.path.basename('config.json')))):
        m_debug['config'] = mdl.config
        m_debug.save()
    return True

def createSymlink(mdl, current_dir=None, symlink_name='Exported', **kwags):
    """
    創建目標路徑的快捷方式到指定目錄
    
    Parameters:
    -----------
    target_path : str
        目標路徑（被連結的目標）
    current_dir : str, optional
        要創建快捷方式的目錄，默認為當前工作目錄
    symlink_name : str, optional
        快捷方式的名稱，默認使用目標路徑的基本名稱
    
    Returns:
    --------
    bool
        創建成功返回True，失敗返回False
    """
    try:
        target_path = mdl.exp_fd
        # 獲取當前工作目錄（如果未指定）
        if current_dir is None: current_dir = '.'
        # 獲取快捷方式名稱（如果未指定）
        if symlink_name is None:    symlink_name = os.path.basename(target_path)
        # 構建快捷方式的完整路徑
        symlink_path = os.path.join(current_dir, symlink_name)
        # 如果已存在同名快捷方式，先刪除
        if os.path.exists(symlink_path):
            os.remove(symlink_path)
        # 創建新的快捷方式
        LOGger.addDebug('symlink_path', symlink_path)
        # os.link(target_path, symlink_path)
        os.symlink(target_path, symlink_path, target_is_directory=True)
        return True
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['createSymlink'])
        return False

def scenario(mdl, config=None, config_file='config.json', projectSheetStamps=None, **kwags):
    if(not mdl.setInitProperty()):
        mdl.addlog('setInitProperty failed', type(mdl), colora=LOGger.FAIL)
        return False
    mdl.addlog('mdl setInitProperty successful', colora=LOGger.OKCYAN)
    if(not mdl.dataPurposeClassification(mdl.source_data)):
        mdl.addlog('dataPurposeClassification failed', type(mdl), colora=LOGger.FAIL)
        return False
    mdl.addlog('mdl dataPurposeClassification successful', colora=LOGger.OKCYAN)
    if(not os.path.isdir(mdl.mdc.get_graph_exp_fd())):   CreateContainer(mdl.mdc.get_graph_exp_fd())
    if(mdl.need_training):
        if(not mdl.fit()):
            mdl.addlog('fit failed', type(mdl), colora=LOGger.FAIL)
            return False
        mdl.addlog('mdl fit successful', colora=LOGger.OKCYAN)
    if(mdl.mdc.model==None):
        mdl.addlog('No model??!??!!?!?', type(mdl), colora=LOGger.FAIL)
        return False
    if(mdl.need_training):
        if(not mdl.mdc.save_config()):
            mdl.addlog('mdl.mdc.save_config failed', colora=LOGger.FAIL)
            return False
        mdl.addlog('mdl.mdc.save_config successful', colora=LOGger.OKCYAN)
    if(not evaluationScenario(mdl, exp_fd=mdl.mdc.get_graph_exp_fd())):
        mdl.addlog('evaluationScenario failed', type(mdl), colora=LOGger.FAIL)
        return False
    mdl.addlog('evaluationScenario successful', colora=LOGger.OKCYAN)
    if(mdl.need_training):
        if(not mdl.mdc.save_config()):
            mdl.addlog('mdl.mdc.save_config failed', colora=LOGger.FAIL)
            return False
        mdl.addlog('mdl.mdc.save_config successful finally', colora=LOGger.OKCYAN)
    if(not mdl.projectRecord()):
        mdl.addlog('mdl projectRecord failed', type(mdl), colora=LOGger.FAIL)
        return False
    mdl.addlog('mdl projectRecord successful', colora=LOGger.OKCYAN)
    if(not saveProjectRecord(mdl, stamps=projectSheetStamps)):
        mdl.addlog('saveProjectRecord failed', type(mdl), colora=LOGger.FAIL)
        return False
    mdl.addlog('saveProjectRecord successful', colora=LOGger.OKCYAN)
    return True

#%%
class abc_EIMS(abc.ABC):
    def __init__(self, config_file='config.json', config=None, exp_fd=None, theme=None, project_exp_fd='.', stamps=None, **kwags):
        """
        project_exp_fd : str
            project_exp_fd 是專案的實驗結果儲存目錄(用來評估各次訓練的績效)
        stamps: list
            stamps 是實驗的標籤(用來評估各次訓練的績效)
        """
        self.config_file = config_file
        self.config = config
        self.exp_fd = exp_fd
        self.supreProperty = {'rewrite': True, 'exp_fd': None, 'theme_layers': [], 'coreClassName':'EIMS_core', 'need_training':False,
                              'auxheader':[], 'tn_ratio':0.85, 'is_valid_data':True, 'stamps':[] }
        self.coreProperty = {'version': '', 'modelFileType':'pkl'}
        self.dataProperty = {'source_data_file':None, 'sht':0, 'sheet':0, 'sheetname':0, 'dataCountLbd':0, 'stratifyLabels':[],
                             'importMethod':"lambda f,**kwags:pd.DataFrame(DFP.import_data(f))"}
        stamps = stamps if(isinstance(stamps, list)) else [] #未來要過繼給mdc
        self.stamps = stamps
        self.source_data = None
        self.coreClassName = 'EIMS_core'
        self.addlog = LOGger.addloger(logfile='')
        self.mdc = None
        self.dfx, self.dfy, self.X_train, self.X_test, self.y_train, self.y_test = pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        self.export = {}
        self.theme = os.path.basename(os.getcwd()) if(not theme) else DFP.parse(theme)
        self.project_exp_fd = project_exp_fd
        self.prm_fn = os.path.join(self.project_exp_fd, '%s.xlsx'%LOGger.stamp_process('', ['records',self.theme], '','','','_'))

    @abc.abstractmethod
    def scenario(self, **kwags):
        pass
    
    def loadConfig(self, config=None, eva_bck='$'):
        config = config if(LOGger.isinstance_not_empty(config, dict)) else self.config
        if(not loadConfigStandard(self, config, eva_bck=eva_bck)):
            return False
        return True
    
    def loadConfigFromFile(self, file=None, eva_bck='$'):
        file = file if(LOGger.isinstance_not_empty(file, str)) else self.config_file
        if(not loadConfigFromFileStandard(self, file, eva_bck=eva_bck)):
            return False
        return True
    
    def setProperty(self, topic):
        return setProperty(self, topic)
    
    def setInputProperty(self):
        return setInputProperty(self, **self.dataProperty)
    
    def setOuputProperty(self):
        return setOuputProperty(self)
        
    def inheritPropToCore(self):
        return inheritPropToCore(self)
        
    def setInitProperty(self):
        """
        setProperty, setOuputProperty, mdc, inheritPropToCore...

        """
        return setInitProperty(self)
    
    def constructOutputDiretory(self):
        return constructOutputDiretory(self)
    
    def createModelFileBasename(self):
        return createModelFileBasename(self)
    
    def get_hyperlink(self):
        return get_hyperlink(self)
        
    def dataPurposeClassification(self, df):
        return dataPurposeClassificationStandard(self, df)
    
    def compiling(self, **kwags):
        return compilingStandard(self, **kwags)
    
    def fit(self, **kwags):
        return fitStandard(self, **kwags)
    
    def evaluation(self, X_data, y_data, **kwags):
        return evaluationStandard(self, X_data, y_data, **kwags)
    
    def projectRecordOfGeneralHeader(self, **kwags):
        return projectRecordOfGeneralHeaderStandard(self)
    
    def projectRecord(self, **kwags):
        return projectRecordStandard(self, **kwags)
    
    def clear(self):
        if(not self.mdc.clear()):
            return False
        return True

class EIMS(abc_EIMS):
    def __init__(self, config_file='', exp_fd=None, figUnfamScoresFigsize=(8,6), figUnfamScoresLayout=(2,1), **kwags):
        super().__init__(config_file=config_file, exp_fd=exp_fd, **kwags)
        self.coreProperty['modelFileType '] = 'h5'

        self.figUnfamScoresFigsize = figUnfamScoresFigsize
        self.figUnfamScoresLayout = figUnfamScoresLayout
    
    def scenario(self, **kwags):
        return scenario(self, **kwags)
    
class EIMS_AutoEncoder(EIMS):
    def __init__(self, config_file='', exp_fd=None, figUnfamScoresFigsize=(14,6), figUnfamScoresLayout=(2,2), **kwags):
        super().__init__(config_file=config_file, exp_fd=exp_fd, 
                         figUnfamScoresFigsize=figUnfamScoresFigsize, 
                         figUnfamScoresLayout=figUnfamScoresLayout, **kwags)
        self.coreProperty['modelFileType '] = 'h5'
        self.supreProperty['coreClassName'] = 'EIMS_AUTOENCODER_core'
        self.coreClassName = 'EIMS_AUTOENCODER_core'
    
    def evaluation(self, X_data, y_data=None, **kwags):
        # y_data is always None
        return evaluationStandard(self, X_data, None, **kwags)
    
    def dataPurposeClassification(self, df):
        return dataPurposeClassificationOnlyCause(self, df)

    def scenario(self, **kwags):
        return scenario(self, **kwags)
    
class EIMS_Sklearn(abc_EIMS):
    def __init__(self, config_file='', exp_fd=None, deepParams=True, **kwags):
        super().__init__(config_file=config_file, exp_fd=exp_fd, **kwags)
        
        self.coreProperty['modelFileType'] = 'pkl'
        self.supreProperty['coreClassName'] = 'EIMS_SKLEARN_core'
        self.supreProperty['deepParams'] = deepParams
        self.coreClassName = 'EIMS_SKLEARN_core'
    
    def scenario(self, **kwags):
        return scenario(self, **kwags)
    
class EIMS_AutoSklearn(abc_EIMS):
    def __init__(self, config_file='', exp_fd=None, deepParams=True, **kwags):
        super().__init__(config_file=config_file, exp_fd=exp_fd, **kwags)
        self.coreProperty['modelFileType'] = 'pkl'
        self.supreProperty['coreClassName'] = 'EIMS_AUTOSKLEARN_core'
        self.supreProperty['deepParams'] = deepParams
        self.coreClassName = 'EIMS_AUTOSKLEARN_core'
    
    def scenario(self, **kwags):
        return scenario(self, **kwags)
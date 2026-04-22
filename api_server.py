from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, validator
from typing import Optional, Dict, Any, List, Tuple
import os
import sys
import json
import math
import time
import queue
import threading
import asyncio
import concurrent.futures
import numpy as np
import pandas as pd
import openpyxl

# 加入當前目錄到 Python 路徑
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from package import LOGger
from src import analyze_engine as ae
m_print = LOGger.addloger(logfile='')
m_logfile = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'api_server_%t.log')
m_addlog = LOGger.addloger(logfile=m_logfile)
m_DataSetPath = ae.m_DataSetPath
m_ExportSetPath = ae.m_ExportSetPath
# 不穩定度 API 預設輸出根目錄（issues/不穩定度計算功能API化.iss）；TODO: 可改讀 config.json
m_instability_api_output_default = r"\\10.1.3.127\ml_home\DataAnalysis\apiOutput"

# 載入配置檔案
def load_config():
    """載入 config.json 配置檔案"""
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['load_config'])
        m_addlog(f"載入配置檔案失敗", stamps=['load_config'], colora=LOGger.FAIL)
        return {
            "Host_IP": "127.0.0.1",
            "Host_Port": 8000,
            "source_dir": "."
        }

# 載入配置
config = load_config()

# 導入 DataAnalysis 模組
import DataAnalysis as DA
from package import dataframeprocedure as DFP
from src import heteroscedastic as het_instability
from src import header_zone_resolver as hz_resolver
from src import plot_correlation_job_registry as pc_job_registry
try:
    from src.discrete_aggregator import aggregate_discrete_data
except Exception:
    aggregate_discrete_data = None

app = FastAPI(
    title="DataAnalysis API",
    description="資料分析 API 服務",
    version="1.0.0"
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
# 掛載靜態文件目錄（CSS、JS）
static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# 模板目錄路徑
template_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'template')

# "filePath": "\\\\10.1.3.129\\EIMSFileData\\ExportSet\\63\\1\\graph\\corr.xlsx",
# "sheet": 0,
# "method": "mic",
# "exp_fd": "tmp\\micCorr",
# "stamps": [
# ],
# "ret": {}

# 請求模型
class FilePathRequest(BaseModel):
    filePath: str
    sheet: Optional[int] = 0
    output_dir: Optional[str] = None

class AnalysisRequest(FilePathRequest):
    xheader: str
    yheader: str
    fixed_values: Optional[Dict[str, Any]] = None
    include_model_prediction: Optional[bool] = True
    ret: Optional[Dict[str, Any]] = {}

class AnalysisViaApiRequest(AnalysisRequest):
    """API 版本的分析請求模型"""
    api_url: Optional[str] = "http://10.1.3.236:5678"
    model_name: Optional[str] = "ACAngle"
    version: Optional[str] = "v0-0-2-0"
    timeout: Optional[int] = 30

class TwoDimensionAnalysisRequest(FilePathRequest):
    xheader1: str
    xheader2: str
    yheader: str
    fixed_values: Optional[Dict[str, Any]] = None
    include_model_prediction: Optional[bool] = True
    ret: Optional[Dict[str, Any]] = {}

class MixedAnalysisRequest(FilePathRequest):
    xheader1: str
    xheader2: str
    yheader: str
    x1_type: str
    x2_type: str
    y_type: str
    fixed_values: Optional[Dict[str, Any]] = None
    include_model_prediction: Optional[bool] = True
    ret: Optional[Dict[str, Any]] = {}

class TwoDimensionAnalysisViaApiRequest(TwoDimensionAnalysisRequest):
    """API 版本的雙維度分析請求模型"""
    api_url: Optional[str] = "http://10.1.3.236:5678"
    model_name: Optional[str] = "ACAngle"
    version: Optional[str] = "v0-0-2-0"
    timeout: Optional[int] = 30
    y_type: Optional[str] = "continuous"  # 用於 analyzeTwoContinuousToY 和 analyzeTwoCategoricalToY

class MixedAnalysisViaApiRequest(MixedAnalysisRequest):
    """API 版本的混合分析請求模型"""
    api_url: Optional[str] = "http://10.1.3.236:5678"
    model_name: Optional[str] = "ACAngle"
    version: Optional[str] = "v0-0-2-0"
    timeout: Optional[int] = 30

class CorrelationRequest(FilePathRequest):
    method: Optional[str] = 'mic'
    exp_fd: Optional[str] = 'micCorr'
    stamps: Optional[List[str]] = None
    ret: Optional[Dict[str, Any]] = {}
    data_config_file: Optional[str] = None
    exclude_headers: Optional[List[str]] = None
    # /plot-correlation：預設非阻塞 job；async_job=False 時維持同步行為
    async_job: Optional[bool] = None
    # 背景 job 逾時（秒）；未傳則用環境變數 DATAANALYSIS_JOB_TIMEOUT_SEC（0 表不限制）
    timeout_sec: Optional[float] = None
    # MIC 進程池 workers；未傳時僅 /plot-correlation 有專用預設，再經 data_analysis.resolve_mic_n_jobs 收斂
    n_jobs: Optional[int] = None
    # 是否先做離散類別聚合（/plot-correlation 專用；預設關閉）
    isAggregate: Optional[bool] = False
    # spill-to-disk：MIC 前欄位分檔、釋放 DataFrame；目前僅後端 pkl
    spill_to_disk: Optional[bool] = False
    spill_backend: Optional[str] = "pkl"
    spill_keep_files: Optional[bool] = False

class DataParsingRequest(FilePathRequest):
    ret: Optional[Dict[str, Any]] = {}
    exp_fd: Optional[str] = 'micCorr'

class DataParsingAndPlotCorrelationRequest(FilePathRequest):
    method: Optional[str] = 'mic'
    exp_fd: Optional[str] = 'micCorr'
    stamps: Optional[List[str]] = None
    ret: Optional[Dict[str, Any]] = {}
    selectHeader: Optional[List[str]] = None
    isAggregate: Optional[bool] = True

class StandardTestsRequest(FilePathRequest):
    column: str
    group_by: Optional[str] = None
    alpha: Optional[float] = 0.05
    output_dir: Optional[str] = None
    use_api: Optional[bool] = False
    plan_id: Optional[str] = None
    seq_no: Optional[str] = None
    # 檢定容入規格（issues/檢定容入規格判斷.iss）；未傳 tol 時行為與舊版一致
    tol: Optional[float] = None
    spec_mode: Optional[str] = "tost"
    confidence: Optional[float] = 0.95
    p_adjust: Optional[str] = "holm"

class PointsRequest(FilePathRequest):
    columns: List[str]                      # 要取出的欄位（2D: [x,y], 3D: [x1,x2,y]）
    fixed_values: Optional[Dict[str, Any]] = None
    sample_n: Optional[int] = 2000
    seed: Optional[int] = 7
    include_extra_columns: Optional[List[str]] = None  # 額外想看詳情的欄位（可選）


class InstabilityBaseRequest(FilePathRequest):
    """不穩定度 API 共用欄位；見 issues/不穩定度計算功能API化.iss"""
    x_col: str
    y_col: str
    layers: Optional[List[int]] = None
    ret: Optional[Dict[str, Any]] = {}
    # layers: None 表三層皆要；否則 1/2/3 — TODO: pydantic 驗證與 heteroscedastic 對齊


class InstabilityComputeRequest(InstabilityBaseRequest):
    """僅 JSON 回傳計算結果，不寫檔、不畫圖（邏輯在 heteroscedastic.compute_instability_payload）"""


class InstabilitySaveRequest(InstabilityBaseRequest):
    """寫檔；output_dir 未給則用 m_instability_api_output_default"""


class InstabilityPlotRequest(InstabilityBaseRequest):
    """繪圖輸出；output_dir 未給則用 m_instability_api_output_default"""


# 安全序列化函數
def safe_serialize(obj, max_depth=3, current_depth=0, _seen=None, **kwags):
    """
    安全地序列化物件，避免循環引用和無窮遞迴
    """
    # 初始化 seen 集合來追蹤已處理的物件
    if _seen is None:
        _seen = set()
        m_print(f'[safe_serialize] 開始序列化，類型: {type(obj).__name__}', colora=LOGger.OKCYAN)
    
    # 檢查循環引用
    obj_id = id(obj)
    if obj_id in _seen:
        m_print(f'[safe_serialize] 檢測到循環引用: {type(obj).__name__}', colora=LOGger.WARNING)
        return f"<circular_reference: {type(obj).__name__}>"
    
    # 防止無窮遞迴
    if current_depth > max_depth:
        m_print(f'[safe_serialize] 達到最大深度限制: {current_depth} > {max_depth}, 類型: {type(obj).__name__}', colora=LOGger.WARNING)
        return f"<max_depth_reached: {type(obj).__name__}>"
    
    # None, bool, str 直接返回
    if obj is None or isinstance(obj, (bool, str)):
        if current_depth == 0:
            m_print(f'[safe_serialize] 簡單類型直接返回: {type(obj).__name__}', colora=LOGger.OKGREEN)
        return obj
    
    # 處理整數類型（包括 numpy 整數類型）
    if isinstance(obj, (int, np.integer)):
        try:
            result = int(obj)
            if current_depth == 0:
                m_print(f'[safe_serialize] 整數類型轉換成功: {type(obj).__name__} -> int', colora=LOGger.OKGREEN)
            return result
        except Exception as e:
            m_print(f'[safe_serialize] 整數轉換失敗: {type(obj).__name__}, 錯誤: {str(e)}', colora=LOGger.FAIL)
            return None
    
    # 處理 float 類型（包括 numpy 浮點類型），檢查是否為 inf、-inf 或 nan
    if isinstance(obj, (float, np.floating)):
        try:
            if math.isnan(obj):
                m_print(f'[safe_serialize] 檢測到 NaN 值 (深度: {current_depth})', colora=LOGger.WARNING)
                return None  # 將 inf、-inf 和 nan 轉換為 None（JSON 中的 null）
            if math.isinf(obj):
                m_print(f'[safe_serialize] 檢測到 Inf 值: {obj} (深度: {current_depth})', colora=LOGger.WARNING)
                return None  # 將 inf、-inf 和 nan 轉換為 None（JSON 中的 null）
            result = float(obj)
            if current_depth == 0:
                m_print(f'[safe_serialize] 浮點數類型轉換成功: {type(obj).__name__} -> float', colora=LOGger.OKGREEN)
            return result
        except (ValueError, TypeError, OverflowError) as e:
            # 如果轉換失敗，返回 None
            m_print(f'[safe_serialize] 浮點數轉換失敗: {type(obj).__name__}, 值: {obj}, 錯誤: {str(e)}', colora=LOGger.FAIL)
            return None
    
    # 將當前物件加入 seen 集合（僅對可變物件）
    if isinstance(obj, (dict, list)):
        _seen.add(obj_id)
    
    try:
        # DataFrame
        if isinstance(obj, pd.DataFrame):
            m_print(f'[safe_serialize] 處理 DataFrame (深度: {current_depth}), 形狀: {obj.shape}', colora=LOGger.OKCYAN)
            try:
                # 安全處理 columns
                columns_list = []
                for col in obj.columns:
                    try:
                        if isinstance(col, (float, np.floating)):
                            if math.isnan(col) or math.isinf(col):
                                columns_list.append("<nan_or_inf_column>")
                            else:
                                columns_list.append(str(float(col)))
                        else:
                            columns_list.append(str(col))
                    except Exception as e:
                        m_print(f'[safe_serialize] DataFrame column 處理失敗: {col}, 錯誤: {str(e)}', colora=LOGger.FAIL)
                        columns_list.append("<column_error>")
                
                # 安全處理 dtypes
                dtypes_dict = {}
                for col, dtype in obj.dtypes.items():
                    try:
                        col_str = str(col) if not (isinstance(col, (float, np.floating)) and (math.isnan(col) or math.isinf(col))) else "<nan_or_inf_column>"
                        dtypes_dict[col_str] = str(dtype)
                    except Exception as e:
                        m_print(f'[safe_serialize] DataFrame dtype 處理失敗: {col}, 錯誤: {str(e)}', colora=LOGger.FAIL)
                        dtypes_dict["<dtype_error>"] = str(dtype)
                
                return {
                    "_type": "DataFrame",
                    "shape": list(obj.shape),
                    "columns": columns_list,
                    "dtypes": dtypes_dict
                }
            except Exception as e:
                m_print(f'[safe_serialize] DataFrame 處理失敗: {str(e)}', colora=LOGger.FAIL)
                raise
        
        # Series
        if isinstance(obj, pd.Series):
            m_print(f'[safe_serialize] 處理 Series (深度: {current_depth}), 形狀: {obj.shape}', colora=LOGger.OKCYAN)
            try:
                return {
                    "_type": "Series",
                    "shape": list(obj.shape),
                    "dtype": str(obj.dtype),
                    "name": str(obj.name) if obj.name is not None else None
                }
            except Exception as e:
                m_print(f'[safe_serialize] Series 處理失敗: {str(e)}', colora=LOGger.FAIL)
                raise
        
        # Numpy array
        if isinstance(obj, np.ndarray):
            m_print(f'[safe_serialize] 處理 ndarray (深度: {current_depth}), 形狀: {obj.shape}, dtype: {obj.dtype}', colora=LOGger.OKCYAN)
            try:
                return {
                    "_type": "ndarray",
                    "shape": list(obj.shape),
                    "dtype": str(obj.dtype)
                }
            except Exception as e:
                m_print(f'[safe_serialize] ndarray 處理失敗: {str(e)}', colora=LOGger.FAIL)
                raise
        
        # 函數
        if callable(obj):
            m_print(f'[safe_serialize] 處理函數: {getattr(obj, "__name__", "unknown")}', colora=LOGger.OKCYAN)
            return f"<function: {getattr(obj, '__name__', 'unknown')}>"
        
        # 字典
        if isinstance(obj, dict):
            m_print(f'[safe_serialize] 處理字典 (深度: {current_depth}), 鍵數量: {len(obj)}', colora=LOGger.OKCYAN)
            result = {}
            for idx, (k, v) in enumerate(obj.items()):
                try:
                    # 處理鍵
                    if isinstance(k, (float, np.floating)):
                        if math.isnan(k) or math.isinf(k):
                            m_print(f'[safe_serialize] 字典鍵 {idx} 為 inf/nan，使用替代鍵名', colora=LOGger.WARNING)
                            key_str = "<nan_or_inf_key>"
                        else:
                            key_str = str(float(k))
                    else:
                        key_str = str(k)
                    
                    # 處理值
                    m_print(f'[safe_serialize] 處理字典鍵 "{key_str}" (深度: {current_depth})', colora=LOGger.OKCYAN)
                    serialized_value = safe_serialize(v, max_depth, current_depth + 1, _seen)
                    result[key_str] = serialized_value
                except Exception as e:
                    m_print(f'[safe_serialize] 字典鍵 "{k}" 處理失敗: {str(e)}', colora=LOGger.FAIL)
                    LOGger.exception_process(e, logfile='', stamps=['safe_serialize'])
                    try:
                        result[str(k)] = f"<error: {str(e)[:30]}>"
                    except:
                        result["<key_error>"] = f"<error: {str(e)[:30]}>"
            return result
        
        # 列表或元組
        if isinstance(obj, (list, tuple)):
            m_print(f'[safe_serialize] 處理 {type(obj).__name__} (深度: {current_depth}), 長度: {len(obj)}', colora=LOGger.OKCYAN)
            try:
                if len(obj) > 50:  # 限制列表長度
                    m_print(f'[safe_serialize] 列表過長 ({len(obj)} > 50)，只處理前5個元素', colora=LOGger.WARNING)
                    return {
                        "_type": type(obj).__name__,
                        "length": len(obj),
                        "sample": [safe_serialize(item, max_depth, current_depth + 1, _seen) for item in list(obj)[:5]]
                    }
                return [safe_serialize(item, max_depth, current_depth + 1, _seen) for item in obj]
            except Exception as e:
                m_print(f'[safe_serialize] {type(obj).__name__} 處理失敗: {str(e)}', colora=LOGger.FAIL)
                LOGger.exception_process(e, logfile='', stamps=['safe_serialize'])
                return f"<{type(obj).__name__}: {len(obj)} items, error: {str(e)[:20]}>"
        
        # 其他物件
        m_print(f'[safe_serialize] 未知類型: {type(obj).__name__} (深度: {current_depth})', colora=LOGger.WARNING)
        return f"<{type(obj).__name__}>"
    
    except Exception as e:
        m_print(f'[safe_serialize] 處理 {type(obj).__name__} 時發生例外 (深度: {current_depth}): {str(e)}', colora=LOGger.FAIL)
        LOGger.exception_process(e, logfile='', stamps=['safe_serialize'])
        return f"<{type(obj).__name__}: error: {str(e)[:50]}>"
    finally:
        # 從 seen 集合中移除（允許在其他分支中再次處理）
        if isinstance(obj, (dict, list)) and obj_id in _seen:
            _seen.discard(obj_id)
            if current_depth == 0:
                m_print(f'[safe_serialize] 序列化完成', colora=LOGger.OKGREEN)

def _normalize_excel_sheet_arg(full_path: str, sheet: Any) -> Any:
    """將 xlsx 的整數分頁索引轉成分頁名稱，避免 import_data 以字串比對時回退第一張。"""
    ext = os.path.splitext(str(full_path))[1].lower()
    if ext not in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return sheet
    if isinstance(sheet, bool) or not isinstance(sheet, int):
        return sheet

    book = None
    try:
        book = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
        sheetnames = list(book.sheetnames or [])
        if not sheetnames:
            raise HTTPException(status_code=400, detail="Excel 沒有任何分頁")

        idx = int(sheet)
        if idx < 0:
            idx = len(sheetnames) + idx
        if idx < 0 or idx >= len(sheetnames):
            raise HTTPException(
                status_code=400,
                detail=f"sheet 索引超出範圍: {sheet}（有效範圍 0~{len(sheetnames)-1}）",
            )

        return sheetnames[idx]
    finally:
        if book is not None:
            try:
                book.close()
            except Exception:
                pass


# 通用資料載入函數
def load_data(file_path: str, sheet: Any = 0):
    """載入資料檔案"""
    try:
        # 如果檔案路徑不是絕對路徑，則從 source_dir 開始尋找
        if not os.path.isabs(file_path):
            # 先檢查是否在當前目錄
            if os.path.exists(file_path):
                full_path = file_path
            else:
                # 從 source_dir 開始尋找
                source_dir = config.get('source_dir', '.')
                full_path = os.path.join(source_dir, file_path)
        else:
            full_path = file_path
        
        if not os.path.exists(full_path):
            # 嘗試在 referenceDirs 中尋找
            reference_dirs = config.get('referenceDirs', [])
            found = False
            for ref_dir in reference_dirs:
                test_path = os.path.join(ref_dir, file_path)
                if os.path.exists(test_path):
                    full_path = test_path
                    found = True
                    break
            
            if not found:
                raise HTTPException(status_code=404, detail=f"檔案不存在: {file_path}")
        
        sheet_for_import = _normalize_excel_sheet_arg(full_path, sheet)
        data = DFP.import_data(full_path, sht=sheet_for_import)
        if data is None:
            raise HTTPException(status_code=400, detail="無法載入資料檔案")
        
        return data
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['load_data'])
        raise HTTPException(status_code=500, detail=f"載入資料時發生錯誤: {str(e)}")


def apply_exclude_headers(data: pd.DataFrame, exclude_headers: Optional[List[str]], log_tag: str) -> pd.DataFrame:
    """排除指定欄位，並保證至少保留一個欄位。"""
    if not isinstance(exclude_headers, list) or not exclude_headers:
        return data

    normalized_excludes = [h for h in exclude_headers if isinstance(h, str) and h.strip()]
    existing_excludes = [h for h in normalized_excludes if h in data.columns]
    missing_excludes = [h for h in normalized_excludes if h not in data.columns]

    if existing_excludes:
        m_print(f'[{log_tag}] 排除欄位: {existing_excludes}', colora=LOGger.OKGREEN)
        data = data.drop(columns=existing_excludes)
    if missing_excludes:
        m_print(f'[{log_tag}] 以下欄位不存在，略過排除: {missing_excludes}', colora=LOGger.WARNING)

    if data.shape[1] == 0:
        raise HTTPException(status_code=400, detail="排除欄位後已無可分析欄位")
    return data


def _resolve_plot_correlation_exp_fd(exp_fd: Optional[str]) -> Tuple[str, str]:
    """回傳 (requested, resolved_absolute)；相對路徑以 API 程序 cwd 為基準轉成絕對路徑。"""
    requested = exp_fd if (exp_fd is not None and str(exp_fd).strip()) else os.path.join("tmp", "micCorr")
    requested = str(requested).strip()
    if os.path.isabs(requested):
        resolved = os.path.normpath(requested)
    else:
        resolved = os.path.abspath(os.path.join(os.getcwd(), requested))
    return requested, resolved


def _effective_plot_job_timeout_sec(request: CorrelationRequest) -> float:
    if request.timeout_sec is not None:
        try:
            return max(0.0, float(request.timeout_sec))
        except (TypeError, ValueError):
            return 0.0
    return pc_job_registry.default_job_timeout_sec()


def _plot_correlation_api_intent_n_jobs(request: CorrelationRequest) -> int:
    """
    僅 /plot-correlation：未傳 n_jobs 時採用 max(min(cpu-1, 4), 1)，再交給 resolve_mic_n_jobs。
    """
    from package.data_analysis import resolve_mic_n_jobs

    c = max(1, (os.cpu_count() or 1))
    default_w = max(min(c - 1, 4), 1)
    if request.n_jobs is None:
        raw = default_w
    else:
        try:
            raw = int(request.n_jobs)
        except (TypeError, ValueError):
            raw = default_w
    return int(resolve_mic_n_jobs(raw))


def _plot_correlation_cpu_gate() -> None:
    """若「可用平行度 cpu-1 < 4」時，待 CPU 較空再執行（有 psutil 則看 cpu_percent）。"""
    c = max(1, (os.cpu_count() or 1))
    if c - 1 >= 4:
        return
    try:
        import psutil  # type: ignore
        for _ in range(200):
            if float(psutil.cpu_percent(interval=0.2)) < 78.0:
                return
    except Exception:
        time.sleep(0.2)


_PLOT_CORR_TASK_QUEUE = queue.Queue()
_PLOT_CORR_WORKER_LOCK = threading.Lock()
_PLOT_CORR_WORKER_STARTED = False


def _plot_correlation_queue_loop() -> None:
    while True:
        try:
            job_id, payload = _PLOT_CORR_TASK_QUEUE.get()
        except Exception:
            continue
        _plot_correlation_cpu_gate()
        _plot_correlation_job_thread_entry(job_id, payload)
        try:
            _PLOT_CORR_TASK_QUEUE.task_done()
        except Exception:
            pass


def _ensure_plot_correlation_queue_worker() -> None:
    global _PLOT_CORR_WORKER_STARTED
    with _PLOT_CORR_WORKER_LOCK:
        if _PLOT_CORR_WORKER_STARTED:
            return
        t = threading.Thread(
            target=_plot_correlation_queue_loop,
            daemon=True,
            name="plot-corr-queue-worker",
        )
        t.start()
        _PLOT_CORR_WORKER_STARTED = True


def _prepare_plot_correlation_dataframe(request: CorrelationRequest) -> pd.DataFrame:
    """載入並套用 data_config_file / exclude_headers（與 /plot-correlation 原本邏輯一致）。"""
    data = load_data(request.filePath, request.sheet)

    if request.data_config_file:
        selected_header = None
        m_print(f'[plot_correlation] 載入配置檔案: {request.data_config_file}', colora=LOGger.OKGREEN)
        data_config = LOGger.load_json(request.data_config_file)
        m_print(f'[plot_correlation] 配置檔案內容: {list(data_config.keys()) if isinstance(data_config, dict) else "非字典格式"}', colora=LOGger.OKGREEN)
        if isinstance(data_config, dict):
            selected_header = data_config.get('selected_header')
            m_print(f'[plot_correlation] selected_header: {selected_header}', colora=LOGger.OKCYAN)
            if not selected_header and 'HEADER' in data_config:
                header_config = data_config.get('HEADER', {})
                m_print(f'[plot_correlation] HEADER 配置: {header_config}', colora=LOGger.OKCYAN)
                if header_config:
                    header_names = []
                    for key, value in header_config.items():
                        if isinstance(value, str) and value:
                            header_names.append(value)
                    if header_names:
                        selected_header = header_names
                        m_print(f'[plot_correlation] 從 HEADER 提取的欄位: {header_names}', colora=LOGger.OKGREEN)
        if isinstance(selected_header, list):
            available_headers = [h for h in selected_header if h in data.columns]
            if available_headers:
                m_print(f'[plot_correlation] 使用選取的欄位: {available_headers}', colora=LOGger.OKGREEN)
                data = data[available_headers]
            else:
                m_print(f'[plot_correlation] 警告: 配置的欄位都不存在於資料中: {selected_header}', colora=LOGger.WARNING)

    data = apply_exclude_headers(data, request.exclude_headers, log_tag='plot_correlation')
    return data


def _apply_plot_correlation_aggregation(
    data: pd.DataFrame,
    request: CorrelationRequest,
    output_dir: str,
) -> pd.DataFrame:
    """依 isAggregate 開關決定是否先做離散類別聚合。"""
    if not bool(getattr(request, "isAggregate", False)):
        return data

    if aggregate_discrete_data is None:
        m_print('[plot_correlation] isAggregate=True 但 discrete_aggregator 無法載入，改用原始資料', colora=LOGger.WARNING)
        return data

    try:
        try:
            # 若當前執行緒已有 event loop（例如 async endpoint），改用子執行緒包 asyncio.run
            asyncio.get_running_loop()
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                fut = executor.submit(asyncio.run, aggregate_discrete_data(data, exp_fd=output_dir))
                aggregated_data, aggregation_results = fut.result()
        except RuntimeError:
            # 無 event loop 時直接 run
            aggregated_data, aggregation_results = asyncio.run(
                aggregate_discrete_data(data, exp_fd=output_dir)
            )

        agg_count = len([
            r for r in (aggregation_results or [])
            if isinstance(r, dict) and r.get('aggregated')
        ])
        m_print(
            f'[plot_correlation] 離散聚合完成，aggregated_columns={agg_count}',
            colora=LOGger.OKCYAN,
        )
        if isinstance(request.ret, dict):
            request.ret['aggregation_results'] = aggregation_results
        return aggregated_data
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['plot_correlation', 'aggregate'])
        m_print(f'[plot_correlation] 離散聚合失敗，改用原始資料: {e}', colora=LOGger.WARNING)
        return data


def _plot_correlation_execute_job(job_id: str, request: CorrelationRequest) -> None:
    """背景執行 MIC 繪圖（單一 worker 內強制同步 plotCorrelation，避免再開內層背景線程）。"""
    timer: Optional[threading.Timer] = None
    try:
        job = pc_job_registry.get_job(job_id)
        if not job:
            return

        if job.status == "cancelled" or job.cancel_event.is_set():
            return

        if job.timeout_sec and float(job.timeout_sec) > 0:
            timer = threading.Timer(
                float(job.timeout_sec),
                lambda jid=job_id: pc_job_registry.try_mark_timeout(jid),
            )
            timer.daemon = True
            timer.start()
            pc_job_registry.set_job_timer(job_id, timer)

        pc_job_registry.mark_running(job_id)
        job = pc_job_registry.get_job(job_id)
        if not job or job.status == "cancelled" or job.cancel_event.is_set():
            return

        try:
            data = _prepare_plot_correlation_dataframe(request)
        except HTTPException as he:
            pc_job_registry.try_mark_failed(job_id, str(he.detail))
            return
        except Exception as e:
            pc_job_registry.try_mark_failed(job_id, str(e))
            return

        job = pc_job_registry.get_job(job_id)
        if not job or job.status in ("failed", "cancelled"):
            return

        os.makedirs(job.resolved_exp_fd, exist_ok=True)
        data = _apply_plot_correlation_aggregation(data, request, job.resolved_exp_fd)

        deadline = None
        if job.timeout_sec and float(job.timeout_sec) > 0:
            deadline = time.monotonic() + float(job.timeout_sec)

        n_eff = int(getattr(job, "effective_n_jobs", 1) or 1)
        m_addlog(
            f"/plot-correlation job_id={job_id} state=run effective_n_jobs={n_eff} path={job.resolved_exp_fd}",
            stamps=['plot_correlation', 'job', str(job_id)],
            colora=LOGger.OKCYAN,
        )
        ret: Dict[str, Any] = request.ret if isinstance(request.ret, dict) else {}
        ok = DA.PlotCorrelation(
            matrix=data,
            method=request.method or "mic",
            exp_fd=job.resolved_exp_fd,
            stamps=request.stamps,
            ret=ret,
            n_jobs=n_eff,
            use_background=False,
            background_async=False,
            resolved_exp_fd=job.resolved_exp_fd,
            job_cancel_event=job.cancel_event,
            job_deadline_monotonic=deadline,
            plot_job_clear_ret=True,
            spill_to_disk=bool(getattr(request, "spill_to_disk", False)),
            spill_backend=str(getattr(request, "spill_backend", "pkl") or "pkl"),
            spill_keep_files=bool(getattr(request, "spill_keep_files", False)),
            plot_job_register_worker_pids=(lambda pids, jid=job_id: pc_job_registry.set_worker_pids(jid, pids)),
            plot_job_note_cancelled_futures=(lambda n, jid=job_id: pc_job_registry.note_cancelled_futures(jid, n)),
            plot_job_note_killed_workers=(lambda n, jid=job_id: pc_job_registry.note_killed_workers(jid, n)),
            plot_job_note_memory_cleared=(lambda jid=job_id: pc_job_registry.note_memory_cleared(jid)),
            plot_job_note_cleanup_error=(lambda msg, jid=job_id: pc_job_registry.note_cleanup_error(jid, msg)),
        )

        jcur = pc_job_registry.get_job(job_id)
        if jcur and jcur.error == "job_timeout":
            return

        if not ok:
            if jcur and jcur.cancel_event.is_set():
                pc_job_registry.try_mark_failed(job_id, "cancelled")
            else:
                pc_job_registry.try_mark_failed(job_id, "PlotCorrelation returned False")
            return

        outs = pc_job_registry.scan_correlation_output_files(job.resolved_exp_fd)
        pc_job_registry.try_mark_success(job_id, outs)
    finally:
        if timer is not None:
            try:
                timer.cancel()
            except Exception:
                pass
        pc_job_registry.clear_job_timer(job_id)


def _plot_correlation_job_thread_entry(job_id: str, payload: Dict[str, Any]) -> None:
    try:
        req = CorrelationRequest(**payload)
    except Exception as e:
        pc_job_registry.try_mark_failed(job_id, f"invalid request payload: {e}")
        return
    _plot_correlation_execute_job(job_id, req)

# 前端模板路由
@app.get("/index.html", response_class=HTMLResponse)
async def index_html():
    """返回首頁（動態注入配置）"""
    index_path = os.path.join(template_dir, 'index.html')
    if os.path.exists(index_path):
        with open(index_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 從 config.json 取得 API 地址
        api_host = config.get("Host_IP", "127.0.0.1")
        api_port = config.get("Host_Port", 8000)
        api_base = f"http://{api_host}:{api_port}"
        
        # 替換硬編碼的 API 地址
        html_content = html_content.replace(
            'http://10.1.3.127:6030',
            api_base
        )
        
        return HTMLResponse(content=html_content)
    raise HTTPException(status_code=404, detail="index.html 不存在")

@app.get("/main.html", response_class=HTMLResponse)
async def main_html():
    """返回主介面頁面（動態注入配置）"""
    main_path = os.path.join(template_dir, 'main.html')
    if os.path.exists(main_path):
        with open(main_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 從 config.json 取得 API 地址
        api_host = config.get("Host_IP", "127.0.0.1")
        api_port = config.get("Host_Port", 8000)
        api_base = f"http://{api_host}:{api_port}"
        
        # 替換硬編碼的 API 地址
        html_content = html_content.replace(
            'http://10.1.3.127:6030',
            api_base
        )
        
        return HTMLResponse(content=html_content)
    raise HTTPException(status_code=404, detail="main.html 不存在")

@app.get("/heteroscedastic.html", response_class=HTMLResponse)
async def heteroscedastic_html():
    """返回不穩定度頁面（動態注入配置）"""
    page_path = os.path.join(template_dir, 'heteroscedastic.html')
    if os.path.exists(page_path):
        with open(page_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        # 從 config.json 取得 API 地址
        api_host = config.get("Host_IP", "127.0.0.1")
        api_port = config.get("Host_Port", 8000)
        api_base = f"http://{api_host}:{api_port}"

        # 替換硬編碼的 API 地址
        html_content = html_content.replace(
            'http://10.1.3.127:6030',
            api_base
        )

        return HTMLResponse(content=html_content)
    raise HTTPException(status_code=404, detail="heteroscedastic.html 不存在")

@app.get("/app.css")
async def app_css():
    """返回 CSS 文件"""
    css_path = os.path.join(static_dir, 'css', 'app.css')
    if os.path.exists(css_path):
        return FileResponse(css_path, media_type="text/css")
    raise HTTPException(status_code=404, detail="app.css 不存在")

@app.get("/app.js")
async def app_js():
    """返回 JavaScript 文件"""
    js_path = os.path.join(static_dir, 'js', 'app.js')
    if os.path.exists(js_path):
        return FileResponse(js_path, media_type="application/javascript")
    raise HTTPException(status_code=404, detail="app.js 不存在")

# API 端點
@app.get("/")
async def root():
    """根路由：返回 API 資訊或重定向到首頁"""
    # 檢查是否有 index.html，如果有則返回 HTML（動態注入配置），否則返回 JSON
    index_path = os.path.join(template_dir, 'index.html')
    if os.path.exists(index_path):
        with open(index_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 從 config.json 取得 API 地址
        api_host = config.get("Host_IP", "127.0.0.1")
        api_port = config.get("Host_Port", 8000)
        api_base = f"http://{api_host}:{api_port}"
        
        # 替換硬編碼的 API 地址
        html_content = html_content.replace(
            'http://10.1.3.127:6030',
            api_base
        )
        
        return HTMLResponse(content=html_content)
    return {
        "message": "DataAnalysis API 服務運行中",
        "config": {
            "host_ip": config.get("Host_IP", "127.0.0.1"),
            "host_port": config.get("Host_Port", 8000),
            "source_dir": config.get("source_dir", "."),
            "reference_dirs": config.get("referenceDirs", [])
        }
    }

@app.get("/config")
async def get_config():
    """取得服務配置資訊"""
    return {
        "host_ip": config.get("Host_IP", "127.0.0.1"),
        "host_port": config.get("Host_Port", 8000),
        "source_dir": config.get("source_dir", "."),
        "reference_dirs": config.get("referenceDirs", []),
        "gpu_memory_limit": config.get("gpuMemoLimit", 10000)
    }


def _resolve_xy_core(
        config_path: Optional[str] = None,
        default_data_path: Optional[str] = None,
        model_config_path: Optional[str] = None,
) -> Dict[str, Any]:
    """issues/動態互動對齊AIAPI.iss 第 4 節：合併設定並解析 x/y。"""
    try:
        merged = hz_resolver.merge_config_sources(
            config_path=config_path,
            default_data_path=default_data_path,
            model_config_path=model_config_path,
        )
    except (ValueError, FileNotFoundError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    ok, msg = hz_resolver.validate_config_dict(merged)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    resolved = hz_resolver.extract_xy_from_config(merged, config_path=config_path)
    ser = safe_serialize(resolved, max_depth=8)
    if isinstance(ser, dict):
        return {"success": True, **ser}
    return {"success": True, "res": ser}


@app.get("/config/resolve-xy")
async def config_resolve_xy(
        config_path: Optional[str] = None,
        default_data_path: Optional[str] = None,
        model_config_path: Optional[str] = None,
):
    """回傳 input_file、x_cols、y_col、xheader_zones、yheader_zones、config_path 等（骨架）。"""
    try:
        return _resolve_xy_core(
            config_path=config_path,
            default_data_path=default_data_path,
            model_config_path=model_config_path,
        )
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['config_resolve_xy'])
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/project-params")
async def project_params(
        config_path: Optional[str] = None,
        default_data_path: Optional[str] = None,
        model_config_path: Optional[str] = None,
):
    """別名，與 GET /config/resolve-xy 相同（issues/動態互動對齊AIAPI.iss）。"""
    try:
        return _resolve_xy_core(
            config_path=config_path,
            default_data_path=default_data_path,
            model_config_path=model_config_path,
        )
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['project_params'])
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/files")
async def list_files(directory: Optional[str] = None):
    """列出 root（source_disk / source_dir）下指定目錄的檔案與子目錄（僅允許瀏覽 root 範圍）"""
    try:
        # 瀏覽根目錄：優先使用 source_disk（config.json: source_disk），否則退回 source_dir
        browse_root = config.get("source_disk") or config.get("source_dir") or "."
        browse_root = str(browse_root).strip() if browse_root is not None else "."
        # Windows drive root: "X:" -> "X:\\"
        if len(browse_root) == 2 and browse_root[1] == ":" and browse_root[0].isalpha():
            browse_root = browse_root + os.sep
        browse_root_abs = os.path.abspath(browse_root)

        source_dir = config.get("source_dir", ".")
        source_dir_abs = os.path.abspath(source_dir)
        # 預設顯示目錄：source_dir 必須位於 browse_root 底下，否則退回 browse_root
        try:
            default_is_under_root = os.path.commonpath([source_dir_abs, browse_root_abs]) == browse_root_abs
        except ValueError:
            default_is_under_root = False
        default_dir_abs = source_dir_abs if default_is_under_root else browse_root_abs

        # directory:
        # - None（未給參數）: 回傳預設目錄（通常是 source_dir）
        # - ""（明確給空字串）: 回傳瀏覽根目錄（browse_root）
        # - 其他: 相對於 browse_root 的子路徑，或絕對路徑（仍需在 browse_root 內）
        if directory is None:
            target_dir_abs = default_dir_abs
        else:
            directory = str(directory).strip()
            if directory == "":
                target_dir_abs = browse_root_abs
            elif os.path.isabs(directory):
                target_dir_abs = os.path.abspath(directory)
            else:
                target_dir_abs = os.path.abspath(os.path.join(browse_root_abs, directory))

        # 僅允許瀏覽 source_dir 範圍（防止跳出根目錄）
        try:
            is_under_source = os.path.commonpath([target_dir_abs, browse_root_abs]) == browse_root_abs
        except ValueError:
            is_under_source = False
        if not is_under_source:
            raise HTTPException(status_code=400, detail=f"僅允許瀏覽 root 底下: {browse_root_abs}")

        if not os.path.exists(target_dir_abs):
            raise HTTPException(status_code=404, detail=f"目錄不存在: {target_dir_abs}")
        if not os.path.isdir(target_dir_abs):
            raise HTTPException(status_code=400, detail=f"不是目錄: {target_dir_abs}")

        directories = []
        files = []
        for item in os.listdir(target_dir_abs):
            item_path = os.path.join(target_dir_abs, item)
            if os.path.isdir(item_path):
                directories.append({
                    "name": item,
                    "path": item_path,
                    "relative_path": os.path.relpath(item_path, browse_root_abs),
                    "modified": os.path.getmtime(item_path)
                })
            elif os.path.isfile(item_path):
                files.append({
                    "name": item,
                    "path": item_path,
                    "relative_path": os.path.relpath(item_path, browse_root_abs),
                    "size": os.path.getsize(item_path),
                    "modified": os.path.getmtime(item_path)
                })

        directories = sorted(directories, key=lambda x: x.get("name", "").lower())
        files = sorted(files, key=lambda x: x.get("name", "").lower())

        relative_directory = os.path.relpath(target_dir_abs, browse_root_abs)
        if relative_directory == ".":
            relative_directory = ""
            parent_relative_directory = None
        else:
            parent_abs = os.path.dirname(target_dir_abs)
            parent_relative_directory = os.path.relpath(parent_abs, browse_root_abs)
            if parent_relative_directory == ".":
                parent_relative_directory = ""

        return {
            # 前端瀏覽用的根目錄（「最上層」）
            "source_dir": browse_root_abs,
            # 原本 config.json 的 source_dir（預設顯示起點），方便 debug
            "source_dir_config": source_dir_abs,
            "directory": target_dir_abs,
            "relative_directory": relative_directory,
            "parent_relative_directory": parent_relative_directory,
            "directories": directories,
            "files": files,
            "count": len(files),
            "dir_count": len(directories),
        }
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['list_files'])
        raise HTTPException(status_code=500, detail=f"列出檔案時發生錯誤: {str(e)}")

@app.get("/search-file")
async def search_file(filename: str):
    """搜尋檔案在配置的目錄中"""
    try:
        search_results = []
        
        # 搜尋來源目錄
        source_dir = config.get("source_dir", ".")
        if os.path.exists(source_dir):
            for root, dirs, files in os.walk(source_dir):
                for file in files:
                    if filename.lower() in file.lower():
                        search_results.append({
                            "name": file,
                            "path": os.path.join(root, file),
                            "directory": root,
                            "found_in": "source_dir"
                        })
        
        # 搜尋參考目錄
        reference_dirs = config.get("referenceDirs", [])
        for ref_dir in reference_dirs:
            if os.path.exists(ref_dir):
                for root, dirs, files in os.walk(ref_dir):
                    for file in files:
                        if filename.lower() in file.lower():
                            search_results.append({
                                "name": file,
                                "path": os.path.join(root, file),
                                "directory": root,
                                "found_in": "reference_dir"
                            })
        
        return {
            "filename": filename,
            "results": search_results,
            "count": len(search_results)
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['search_file'])
        raise HTTPException(status_code=500, detail=f"搜尋檔案時發生錯誤: {str(e)}")

@app.post("/plot-correlation")
async def plot_correlation(request: CorrelationRequest):
    """繪製相關性圖表。預設非阻塞：立即回傳 job_id；async_job=false 時同步完成（舊行為）。"""
    try:
        requested_dir, resolved_dir = _resolve_plot_correlation_exp_fd(
            request.exp_fd or os.path.join("tmp", "micCorr")
        )

        use_async = True if request.async_job is None else bool(request.async_job)
        eff_nj = _plot_correlation_api_intent_n_jobs(request)
        if not use_async:
            data = _prepare_plot_correlation_dataframe(request)
            data = _apply_plot_correlation_aggregation(data, request, resolved_dir)
            result = DA.PlotCorrelation(
                matrix=data,
                method=request.method,
                exp_fd=resolved_dir,
                stamps=request.stamps,
                ret=request.ret,
                n_jobs=eff_nj,
                use_background=False,
                background_async=False,
                resolved_exp_fd=resolved_dir,
                spill_to_disk=bool(getattr(request, "spill_to_disk", False)),
                spill_backend=str(getattr(request, "spill_backend", "pkl") or "pkl"),
                spill_keep_files=bool(getattr(request, "spill_keep_files", False)),
            )
            return {
                "success": result,
                "accepted": False,
                "message": "相關性圖表繪製完成" if result else "相關性圖表繪製失敗",
                "output_directory": resolved_dir,
                "requested_output_directory": requested_dir,
                "resolved_output_directory": resolved_dir,
                "effective_n_jobs": eff_nj,
            }

        timeout_applied = _effective_plot_job_timeout_sec(request)
        job = pc_job_registry.create_job(
            requested_dir,
            resolved_dir,
            effective_n_jobs=eff_nj,
            method=str(request.method or "mic"),
            file_path=str(request.filePath),
            sheet=request.sheet,
            timeout_sec=timeout_applied,
        )
        try:
            payload = request.model_dump()
        except AttributeError:
            payload = request.dict()

        m_addlog(
            f"/plot-correlation 已入列 job_id={job.job_id} effective_n_jobs={eff_nj} base={resolved_dir} workdir={job.resolved_exp_fd}",
            stamps=['plot_correlation', 'job', str(job.job_id)],
            colora=LOGger.OKCYAN,
        )
        _ensure_plot_correlation_queue_worker()
        _PLOT_CORR_TASK_QUEUE.put((job.job_id, payload))

        return {
            "success": True,
            "accepted": True,
            "job_id": job.job_id,
            "message": "已入列背景繪圖佇列，請以 job_id 查詢 /plot-correlation/status/{job_id}",
            "requested_output_directory": requested_dir,
            "resolved_output_directory": job.resolved_exp_fd,
            "output_directory": job.resolved_exp_fd,
            "timeout_sec": timeout_applied,
            "effective_n_jobs": eff_nj,
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['plot_correlation'])
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/plot-correlation/status/{job_id}")
async def plot_correlation_status(job_id: str):
    """查詢 /plot-correlation 背景任務狀態與輸出檔案。"""
    job = pc_job_registry.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job_id 不存在")
    cleanup = pc_job_registry.get_cleanup_snapshot(job_id)
    return {
        "job_id": job.job_id,
        "status": job.status,
        "error": job.error,
        "requested_output_directory": job.requested_exp_fd,
        "resolved_output_directory": job.resolved_exp_fd,
        "output_files": list(job.output_files),
        "started_at": job.started_at,
        "ended_at": job.ended_at,
        "effective_n_jobs": int(getattr(job, "effective_n_jobs", 1) or 1),
        "cleanup": cleanup,
    }


@app.get("/plot-correlation/jobs")
async def plot_correlation_jobs_list():
    """列出 in-memory 中的 plot-correlation 任務摘要。"""
    return {
        "success": True,
        "jobs": pc_job_registry.list_all_job_summaries(),
    }


@app.post("/plot-correlation/jobs/{job_id}/cancel")
async def plot_correlation_cancel(job_id: str):
    """中止背景任務（最佳努力；MIC 計算中無法強制中斷）。"""
    ok = pc_job_registry.request_cancel(job_id)
    if not ok:
        raise HTTPException(status_code=404, detail="job_id 不存在")
    job = pc_job_registry.get_job(job_id)
    cleanup = pc_job_registry.get_cleanup_snapshot(job_id)
    return {
        "success": True,
        "job_id": job_id,
        "status": getattr(job, "status", "unknown"),
        "message": "已送出取消",
        "cleanup": cleanup,
    }

@app.post("/calculate-correlation")
async def calculate_correlation(request: CorrelationRequest):
    """計算相關性"""
    try:
        data = load_data(request.filePath, request.sheet)
        if request.data_config_file:
            selected_header = None
            m_print(f'[calculate_correlation] 載入配置檔案: {request.data_config_file}', colora=LOGger.OKCYAN)
            data_config = LOGger.load_json(request.data_config_file)
            m_print(f'[calculate_correlation] 配置檔案內容: {list(data_config.keys()) if isinstance(data_config, dict) else "非字典格式"}', colora=LOGger.OKCYAN)
            if isinstance(data_config, dict):
                # 優先使用 selected_header
                selected_header = data_config.get('selected_header')
                m_print(f'[calculate_correlation] selected_header: {selected_header}', colora=LOGger.OKCYAN)
                # 如果沒有 selected_header，從 HEADER 配置中提取
                if not selected_header and 'HEADER' in data_config:
                    header_config = data_config.get('HEADER', {})
                    m_print(f'[calculate_correlation] HEADER 配置: {header_config}', colora=LOGger.OKCYAN)
                    if header_config:
                        header_names = []
                        for key, value in header_config.items():
                            if isinstance(value, str) and value:
                                header_names.append(value)
                        if header_names:
                            selected_header = header_names
                            m_print(f'[calculate_correlation] 從 HEADER 提取的欄位: {header_names}', colora=LOGger.OKGREEN)
            if isinstance(selected_header, list):
                # 只保留存在於資料中的欄位
                available_headers = [h for h in selected_header if h in data.columns]
                if available_headers:
                    m_print(f'[calculate_correlation] 使用選取的欄位: {available_headers}', colora=LOGger.OKGREEN)
                    data = data[available_headers]
                else:
                    m_print(f'[calculate_correlation] 警告: 配置的欄位都不存在於資料中: {selected_header}', colora=LOGger.WARNING)
        data = apply_exclude_headers(data, request.exclude_headers, log_tag='calculate_correlation')

        # 設定預設輸出目錄
        exp_fd = request.exp_fd or os.path.join('tmp', 'micCorr')
        
        result = DA.CalculateCorrelation(
            matrix=data,
            method=request.method,
            exp_fd=exp_fd,
            stamps=request.stamps,
            ret=request.ret
        )
        
        # 使用安全序列化
        m_print('使用安全序列化', colora=LOGger.WARNING)
        serializable_ret = safe_serialize(request.ret)
        
        response_dict = {
            "success": result,
            "message": "相關性計算完成" if result else "相關性計算失敗",
            "output_directory": str(exp_fd),  # 確保是字符串
            "res": serializable_ret
        }
        
        # 驗證響應字典中的所有值都是可序列化的
        try:
            import json
            json.dumps(response_dict)  # 測試是否可以序列化
            m_print(f'[calculate_correlation] 響應字典序列化測試通過', colora=LOGger.OKGREEN)
        except (ValueError, TypeError) as e:
            m_print(f'[calculate_correlation] 響應字典序列化測試失敗: {str(e)}', colora=LOGger.FAIL)
            # 如果失敗，再次對整個響應進行安全序列化
            response_dict = safe_serialize(response_dict)
            m_print(f'[calculate_correlation] 已對響應進行安全序列化', colora=LOGger.WARNING)
        
        return response_dict
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['calculate_correlation'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/data-parsing")
async def data_parsing(request: DataParsingRequest):
    """資料解析"""
    try:
        result = DA.DataParsingFromFile(request.filePath, ret=request.ret, sht=request.sheet)
        
        # 使用安全序列化
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "資料解析完成" if result else "資料解析失敗",
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['data_parsing'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/dppc")
async def data_parsing_and_plot_correlation(request: DataParsingAndPlotCorrelationRequest):
    """資料解析並繪製相關性圖表"""
    try:
        print(f"[API] /dppc 開始處理，filePath: {request.filePath}, stamps: {request.stamps}, exp_fd: {request.exp_fd}")
        
        result = DA.DataParsingAndPlotCorrelation(
            request.filePath, 
            ret=request.ret, 
            selectHeader=request.selectHeader,
            exp_fd=request.exp_fd,
            stamps=request.stamps,
            method=request.method,
            isAggregate=request.isAggregate
        )
        
        m_addlog(f"DataParsingAndPlotCorrelation 執行完成，result: {result}", stamps=['data_parsing_and_plot_correlation'], colora=LOGger.OKCYAN)
        
        if not result:
            m_addlog(f"DataParsingAndPlotCorrelation 返回 False，但繼續處理序列化", stamps=['data_parsing_and_plot_correlation'], colora=LOGger.WARNING)
        
        # 使用安全序列化函數處理 ret
        try:
            serializable_ret = safe_serialize(request.ret)
            m_addlog(f"序列化完成", stamps=['data_parsing_and_plot_correlation'], colora=LOGger.OKCYAN)
        except Exception as serialize_error:
            m_addlog(f"序列化失敗: {str(serialize_error)}", stamps=['data_parsing_and_plot_correlation'], colora=LOGger.FAIL)
            # 如果序列化失敗，至少返回基本的結果
            serializable_ret = {"error": "序列化失敗", "message": str(serialize_error)}
        
        return {
            "success": result,
            "message": "資料解析並繪製相關性圖表完成" if result else "資料解析並繪製相關性圖表失敗",
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['data_parsing_and_plot_correlation'])
        m_addlog(f"/dppc 端點錯誤", stamps=['data_parsing_and_plot_correlation'], colora=LOGger.FAIL)
        # 確保錯誤訊息能正確返回
        raise HTTPException(status_code=500, detail=f"處理失敗: {str(e)}")

@app.post("/analyze-continuous-to-continuous")
async def analyze_continuous_to_continuous(request: AnalysisRequest):
    """連續變數對連續變數分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeContinuousToContinuous(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            mdc=None,
            output_dir=request.output_dir
        )
        
        return {
            "success": result,
            "message": "連續變數對連續變數分析完成" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_continuous_to_continuous'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-categorical-to-continuous")
async def analyze_categorical_to_continuous(request: AnalysisRequest):
    """類別變數對連續變數分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeCategoricalToContinuous(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            mdc=None,
            output_dir=request.output_dir
        )
        
        return {
            "success": result,
            "message": "類別變數對連續變數分析完成" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_continuous_to_categorical'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-continuous-to-categorical")
async def analyze_continuous_to_categorical(request: AnalysisRequest):
    """連續變數對類別變數分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeContinuousToCategorical(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            ret=request.ret,
            mdc=None,
            output_dir=request.output_dir
        )
        
        return {
            "success": result,
            "message": "連續變數對類別變數分析完成" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_categorical_to_categorical'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-categorical-to-categorical")
async def analyze_categorical_to_categorical(request: AnalysisRequest):
    """類別變數對類別變數分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeCategoricalToCategorical(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            ret=request.ret,
            mdc=None
        )
        
        return {
            "success": result,
            "message": "類別變數對類別變數分析完成" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_categorical_to_categorical'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-continuous-to-continuous-via-api")
async def analyze_continuous_to_continuous_via_api(request: AnalysisViaApiRequest):
    """連續變數對連續變數分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeContinuousToContinuous_via_api(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "連續變數對連續變數分析完成（API 版本）" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader,
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_continuous_to_continuous_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-categorical-to-continuous-via-api")
async def analyze_categorical_to_continuous_via_api(request: AnalysisViaApiRequest):
    """類別變數對連續變數分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeCategoricalToContinuous_via_api(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "類別變數對連續變數分析完成（API 版本）" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader,
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_categorical_to_continuous_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-continuous-to-categorical-via-api")
async def analyze_continuous_to_categorical_via_api(request: AnalysisViaApiRequest):
    """連續變數對類別變數分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeContinuousToCategorical_via_api(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "連續變數對類別變數分析完成（API 版本）" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader,
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_continuous_to_categorical_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-categorical-to-categorical-via-api")
async def analyze_categorical_to_categorical_via_api(request: AnalysisViaApiRequest):
    """類別變數對類別變數分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeCategoricalToCategorical_via_api(
            data=data,
            xheader=request.xheader,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "類別變數對類別變數分析完成（API 版本）" if result else "分析失敗",
            "xheader": request.xheader,
            "yheader": request.yheader,
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_categorical_to_categorical_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/two-dimension-analysis")
async def two_dimension_analysis(request: TwoDimensionAnalysisRequest):
    """二維度分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.twoDimensionAnalysis(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret
        )
        
        return {
            "success": result,
            "message": "二維度分析完成" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['two_dimension_analysis'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-two-continuous-to-y")
async def analyze_two_continuous_to_y(request: TwoDimensionAnalysisRequest):
    """兩個連續變數對 Y 分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeTwoContinuousToY(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            y_type="continuous",  # 預設為連續變數
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret
        )
        
        return {
            "success": result,
            "message": "兩個連續變數對 Y 分析完成" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_two_continuous_to_y'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-mixed-to-y")
async def analyze_mixed_to_y(request: MixedAnalysisRequest):
    """混合變數對 Y 分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeMixedToY(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            x1_type=request.x1_type,
            x2_type=request.x2_type,
            y_type=request.y_type,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret
        )
        
        return {
            "success": result,
            "message": "混合變數對 Y 分析完成" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader,
            "x1_type": request.x1_type,
            "x2_type": request.x2_type,
            "y_type": request.y_type
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_mixed_to_y'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-two-categorical-to-y")
async def analyze_two_categorical_to_y(request: TwoDimensionAnalysisRequest):
    """兩個類別變數對 Y 分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeTwoCategoricalToY(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            y_type="categorical",  # 預設為類別變數
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret
        )
        
        return {
            "success": result,
            "message": "兩個類別變數對 Y 分析完成" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_two_categorical_to_y'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-two-continuous-to-y-via-api")
async def analyze_two_continuous_to_y_via_api(request: TwoDimensionAnalysisViaApiRequest):
    """兩個連續變數對 Y 分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeTwoContinuousToY_via_api(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            y_type=request.y_type or "continuous",
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "兩個連續變數對 Y 分析完成（API 版本）" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader,
            "y_type": request.y_type or "continuous",
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_two_continuous_to_y_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-mixed-to-y-via-api")
async def analyze_mixed_to_y_via_api(request: MixedAnalysisViaApiRequest):
    """混合變數對 Y 分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeMixedToY_via_api(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            x1_type=request.x1_type,
            x2_type=request.x2_type,
            y_type=request.y_type,
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "混合變數對 Y 分析完成（API 版本）" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader,
            "x1_type": request.x1_type,
            "x2_type": request.x2_type,
            "y_type": request.y_type,
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_mixed_to_y_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze-two-categorical-to-y-via-api")
async def analyze_two_categorical_to_y_via_api(request: TwoDimensionAnalysisViaApiRequest):
    """兩個類別變數對 Y 分析（API 版本）"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        result = DA.analyzeTwoCategoricalToY_via_api(
            data=data,
            xheader1=request.xheader1,
            xheader2=request.xheader2,
            yheader=request.yheader,
            y_type=request.y_type or "categorical",
            fixed_values=request.fixed_values,
            include_model_prediction=request.include_model_prediction,
            ret=request.ret,
            output_dir=request.output_dir,
            api_url=request.api_url,
            model_name=request.model_name,
            version=request.version,
            timeout=request.timeout
        )
        
        # 使用安全序列化處理 ret
        serializable_ret = safe_serialize(request.ret)
        
        return {
            "success": result,
            "message": "兩個類別變數對 Y 分析完成（API 版本）" if result else "分析失敗",
            "xheader1": request.xheader1,
            "xheader2": request.xheader2,
            "yheader": request.yheader,
            "y_type": request.y_type or "categorical",
            "api_url": request.api_url,
            "model_name": request.model_name,
            "version": request.version,
            "res": serializable_ret
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['analyze_two_categorical_to_y_via_api'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/standard-tests")
async def standard_tests(request: StandardTestsRequest):
    """標準檢定分析"""
    try:
        data = load_data(request.filePath, request.sheet)
        
        # 建立 handler 物件
        class Handler:
            def __init__(self, **kwargs):
                for key, value in kwargs.items():
                    setattr(self, key, value)
        
        handler = Handler(
            file_path=request.filePath,
            column=request.column,
            group_by=request.group_by,
            sheet=request.sheet,
            alpha=request.alpha,
            output_dir=request.output_dir or os.path.join('tmp', 'distribution_results'),
            use_api=request.use_api,
            plan_id=request.plan_id,
            seq_no=request.seq_no,
            tol=request.tol,
            spec_mode=request.spec_mode,
            confidence=request.confidence,
            p_adjust=request.p_adjust,
        )
        
        # 執行標準檢定
        DA.standardTests(handler)
        
        return {
            "success": True,
            "message": "標準檢定分析完成",
            "column": request.column,
            "group_by": request.group_by,
            "output_directory": handler.output_dir
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['standard_tests'])
        raise HTTPException(status_code=500, detail=str(e))


def _instability_resolve_output_dir(request: InstabilityBaseRequest) -> str:
    # TODO: 與 config 合併、子目錄命名（job_id 等）
    return request.output_dir or m_instability_api_output_default


def _plotly_scalar(v):
    if v is None:
        return None
    if isinstance(v, (pd.Timestamp,)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, (np.integer, int)):
        try:
            return int(v)
        except Exception:
            return None
    if isinstance(v, (np.floating, float)):
        try:
            vf = float(v)
            return vf if np.isfinite(vf) else None
        except Exception:
            return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    return v


def _plotly_as_float(v):
    try:
        vf = float(v)
        return vf if np.isfinite(vf) else None
    except Exception:
        return None


def _build_instability_plotly_payload(
    df: pd.DataFrame,
    payload: Dict[str, Any],
    x_col: str,
    y_col: str,
    extra: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    extra = extra or {}
    try:
        sample_n = int(extra.get("plotly_sample_n", 3000))
    except Exception:
        sample_n = 3000
    sample_n = max(1, min(sample_n, 20000))

    try:
        seed = int(extra.get("plotly_seed", 7))
    except Exception:
        seed = 7

    if x_col not in df.columns or y_col not in df.columns:
        return None

    sub = df[[x_col, y_col]].copy()
    sub = sub.dropna(subset=[x_col, y_col])
    if sub.empty:
        return None

    # y 軸以數值為主；無法轉數值者略過
    y_numeric = pd.to_numeric(sub[y_col], errors="coerce")
    sub = sub.loc[y_numeric.notna()].copy()
    if sub.empty:
        return None
    sub[y_col] = y_numeric.loc[sub.index].astype(float)

    if len(sub) > sample_n:
        rng = np.random.default_rng(seed)
        idx = rng.choice(len(sub), size=sample_n, replace=False)
        sub = sub.iloc[idx]

    x_vals = []
    y_vals = []
    for xv, yv in zip(sub[x_col].tolist(), sub[y_col].tolist()):
        xs = _plotly_scalar(xv)
        ys = _plotly_as_float(yv)
        if ys is None:
            continue
        x_vals.append(xs)
        y_vals.append(ys)
    if not y_vals:
        return None

    def _trace_axis_name(axis_prefix: str, row_idx: int) -> str:
        return axis_prefix if row_idx == 1 else f"{axis_prefix}{row_idx}"

    traces: List[Dict[str, Any]] = []
    row_titles: List[str] = [f"{x_col} vs {y_col}"]
    row_specs: List[Dict[str, Any]] = [{"y_title": y_col}]

    traces.append({
        "type": "scattergl",
        "mode": "markers",
        "name": f"{x_col} vs {y_col}",
        "x": x_vals,
        "y": y_vals,
        "xaxis": _trace_axis_name("x", 1),
        "yaxis": _trace_axis_name("y", 1),
        "marker": {
            "size": 5,
            "color": "#66e3ff",
            "opacity": 0.42,
            "line": {"width": 0},
        },
        "hovertemplate": f"{x_col}=%{{x}}<br>{y_col}=%{{y}}<extra></extra>",
    })

    layer_preview = payload.get("layers_table_preview")
    if isinstance(layer_preview, list) and len(layer_preview) > 1:
        ordered = [r for r in layer_preview if isinstance(r, dict) and (x_col in r)]
        x_numeric = bool(ordered) and all(_plotly_as_float(r.get(x_col)) is not None for r in ordered)
        if x_numeric:
            ordered = sorted(ordered, key=lambda r: _plotly_as_float(r.get(x_col)))

        grouped_candidates = [
            ("Layer1 sigma (preview)", "Layer1 sigma", [("layer1_instability_sigma", "Layer1 sigma", "#7dd3fc")]),
            ("Layer2 sigma/tol (preview)", "Layer2 sigma/tol", [("layer2_instability_sigma_vs_tol", "Layer2 sigma/tol", "#fbbf24")]),
            ("Layer3 score/stability (preview)", "Layer3 metric", [
                ("layer3_instability_score", "Layer3 score", "#f87171"),
                ("layer3_stability_score", "Layer3 stability", "#86efac"),
            ]),
        ]

        for panel_title, y_title, candidates in grouped_candidates:
            panel_traces: List[Dict[str, Any]] = []
            for col, name, color in candidates:
                points = []
                for row in ordered:
                    xv = _plotly_scalar(row.get(x_col))
                    yv = _plotly_as_float(row.get(col))
                    if yv is None:
                        continue
                    points.append((xv, yv))
                if len(points) < 2:
                    continue
                panel_traces.append({
                    "type": "scatter",
                    "mode": "lines+markers",
                    "name": f"{name} (preview)",
                    "x": [p[0] for p in points],
                    "y": [p[1] for p in points],
                    "marker": {"size": 5, "color": color},
                    "line": {"width": 2, "color": color},
                    "hovertemplate": f"{x_col}=%{{x}}<br>{name}=%{{y}}<extra></extra>",
                })
            if not panel_traces:
                continue

            row_idx = len(row_specs) + 1
            for t in panel_traces:
                t["xaxis"] = _trace_axis_name("x", row_idx)
                t["yaxis"] = _trace_axis_name("y", row_idx)
                traces.append(t)
            row_specs.append({"y_title": y_title})
            row_titles.append(panel_title)

    row_count = len(row_specs)
    gap = 0.04 if row_count > 1 else 0.0
    usable = max(0.1, 1.0 - gap * (row_count - 1))
    panel_height = usable / row_count

    layout: Dict[str, Any] = {
        "height": max(560, 230 * row_count),
        "margin": {"l": 68, "r": 68, "t": 96, "b": 64},
        "paper_bgcolor": "rgba(0,0,0,0)",
        "plot_bgcolor": "rgba(7,10,16,.45)",
        "font": {"color": "#dbeafe"},
        "legend": {"orientation": "h", "y": 1.06, "x": 0},
        "hovermode": "closest",
    }

    annotations: List[Dict[str, Any]] = []
    for idx, spec in enumerate(row_specs):
        top = 1.0 - idx * (panel_height + gap)
        bottom = top - panel_height

        xaxis_key = "xaxis" if idx == 0 else f"xaxis{idx + 1}"
        yaxis_key = "yaxis" if idx == 0 else f"yaxis{idx + 1}"
        show_bottom_ticks = (idx == row_count - 1)

        layout[xaxis_key] = {
            "title": x_col if show_bottom_ticks else "",
            "domain": [0.0, 1.0],
            "anchor": _trace_axis_name("y", idx + 1),
            "gridcolor": "rgba(29,42,85,.45)",
            "zerolinecolor": "rgba(29,42,85,.45)",
            "showticklabels": show_bottom_ticks,
            "automargin": True,
        }
        layout[yaxis_key] = {
            "title": spec["y_title"],
            "domain": [max(0.0, bottom), min(1.0, top)],
            "anchor": _trace_axis_name("x", idx + 1),
            "gridcolor": "rgba(29,42,85,.45)",
            "zerolinecolor": "rgba(29,42,85,.45)",
            "rangemode": "tozero",
            "automargin": True,
        }

        annotations.append({
            "text": row_titles[idx],
            "xref": "paper",
            "yref": "paper",
            "x": 0.0,
            "y": min(1.0, top + 0.012),
            "xanchor": "left",
            "yanchor": "bottom",
            "showarrow": False,
            "font": {"size": 12, "color": "#dbeafe"},
        })

    if annotations:
        layout["annotations"] = annotations

    config = {
        "responsive": True,
        "displaylogo": False,
        "modeBarButtonsToRemove": ["lasso2d"],
    }
    return {
        "data": traces,
        "layout": layout,
        "config": config,
        "meta": {
            "point_count": len(y_vals),
            "sample_n": sample_n,
            "x_col": x_col,
            "y_col": y_col,
            "panel_count": row_count,
        },
    }


def _build_instability_view_payload(df: pd.DataFrame, request: InstabilityBaseRequest) -> Dict[str, Any]:
    payload = het_instability.compute_instability_payload(
        df,
        request.x_col,
        request.y_col,
        layers=request.layers,
        extra=request.ret,
    )
    plotly_payload = _build_instability_plotly_payload(
        df=df,
        payload=payload,
        x_col=request.x_col,
        y_col=request.y_col,
        extra=request.ret,
    )
    if isinstance(plotly_payload, dict):
        payload["plotly"] = plotly_payload
    return payload


@app.post("/instability/compute")
async def instability_compute(request: InstabilityComputeRequest):
    """
    不穩定度計算，主體回 JSON（issues/不穩定度計算功能API化.iss）。
    TODO: 將 ret 與 heteroscedastic 參數（離散閾值、norm、tol、density_shrink 等）對齊。
    """
    try:
        df = load_data(request.filePath, request.sheet)
        payload = _build_instability_view_payload(df, request)
        return {
            "success": True,
            "message": "不穩定度計算（骨架）",
            "filePath": request.filePath,
            "x_col": request.x_col,
            "y_col": request.y_col,
            "layers": request.layers,
            "res": safe_serialize(payload, max_depth=8),
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['instability_compute'])
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/instability/save")
async def instability_save(request: InstabilitySaveRequest):
    """不穩定度結果寫檔（與 compute / plot 分離）。"""
    try:
        df = load_data(request.filePath, request.sheet)
        out_dir = _instability_resolve_output_dir(request)
        # TODO: os.makedirs(out_dir, exist_ok=True)
        saved = het_instability.save_instability_to_disk(
            df,
            request.x_col,
            request.y_col,
            output_base_dir=out_dir,
            layers=request.layers,
            extra=request.ret,
        )
        view_payload = _build_instability_view_payload(df, request)
        saved["layers_table_preview"] = view_payload.get("layers_table_preview", [])
        saved["layer_columns"] = view_payload.get("layer_columns", [])
        saved["notes"] = view_payload.get("notes", [])
        if isinstance(view_payload.get("plotly"), dict):
            saved["plotly"] = view_payload["plotly"]
        return {
            "success": True,
            "message": "不穩定度存檔（骨架）",
            "output_dir": out_dir,
            "res": safe_serialize(saved, max_depth=8),
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['instability_save'])
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/instability/plot")
async def instability_plot(request: InstabilityPlotRequest):
    """不穩定度繪圖（與 compute / save 分離）。"""
    try:
        df = load_data(request.filePath, request.sheet)
        out_dir = _instability_resolve_output_dir(request)
        plotted = het_instability.plot_instability_figures(
            df,
            request.x_col,
            request.y_col,
            output_base_dir=out_dir,
            layers=request.layers,
            extra=request.ret,
        )
        view_payload = _build_instability_view_payload(df, request)
        plotted["layers_table_preview"] = view_payload.get("layers_table_preview", [])
        plotted["layer_columns"] = view_payload.get("layer_columns", [])
        plotted["notes"] = view_payload.get("notes", [])
        if isinstance(view_payload.get("plotly"), dict):
            plotted["plotly"] = view_payload["plotly"]
        return {
            "success": True,
            "message": "不穩定度繪圖（骨架）",
            "output_dir": out_dir,
            "res": safe_serialize(plotted, max_depth=8),
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['instability_plot'])
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/data/points")
async def get_points(request: PointsRequest):
    """
    回傳前端繪圖用的資料點（2D/3D），支援：
    - 指定 columns（必要）
    - fixed_values 條件過濾（可選）
    - sample_n 抽樣（避免點太多）
    - include_extra_columns 額外欄位，供 hover/click 顯示更多特徵
    """
    try:
        df = load_data(request.filePath, request.sheet)

        cols = list(dict.fromkeys(request.columns + (request.include_extra_columns or [])))
        missing = [c for c in cols if c not in df.columns]
        if missing:
            raise HTTPException(status_code=400, detail=f"欄位不存在: {missing[:10]}")

        sub = df[cols].copy()

        # fixed_values filter
        if isinstance(request.fixed_values, dict) and request.fixed_values:
            for k, v in request.fixed_values.items():
                if k in sub.columns:
                    sub = sub[sub[k] == v]

        # drop rows with NaN in required columns
        required = request.columns
        sub = sub.dropna(subset=required)

        # sampling
        n = int(request.sample_n or 2000)
        if n > 0 and len(sub) > n:
            rng = np.random.default_rng(int(request.seed or 7))
            idx = rng.choice(len(sub), size=n, replace=False)
            sub = sub.iloc[idx]

        # to jsonable
        # - 將 numpy/pandas 型別轉成 python 基本型別，避免 JSON 序列化問題
        rows = []
        for i, row in sub.iterrows():
            d = {}
            for c in cols:
                val = row[c]
                if pd.isna(val):
                    d[c] = None
                elif isinstance(val, (np.integer,)):
                    d[c] = int(val)
                elif isinstance(val, (np.floating,)):
                    d[c] = float(val)
                else:
                    d[c] = str(val) if isinstance(val, (pd.Timestamp,)) else val
            d["_index"] = int(i) if isinstance(i, (int, np.integer)) else str(i)
            rows.append(d)

        return {
            "success": True,
            "count": len(rows),
            "required_columns": request.columns,
            "columns": cols,
            "rows": rows
        }
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['get_points'])
        raise HTTPException(status_code=500, detail=str(e))


class EvaluationFitRequest(BaseModel):
    plan_id: int = Field(..., gt=0, description="計劃序號，必須為正整數")
    seq_no: int = Field(1, gt=0, description="序列號，必須為正整數，預設為 1")
    stamps: Optional[List[str]] = Field(None, description="額外的標記列表")
    output_dir: Optional[str] = Field(None, description="輸出目錄，可選")
    provider: Optional[str] = Field(None, description="LLM provider，預設為None，後面系統自動帶")
    model: Optional[str] = Field("remote8b", description="模型名稱，預設為 'remote8b'")
    user_prompt_template: Optional[str] = Field(None, description="用戶提示模板，可選")
    system_prompt: Optional[str] = Field(None, description="系統提示模板，可選")
    prompt_temperature: Optional[float] = Field(None, description="提示溫度，可選")
    personality: Optional[str] = Field(None, description="LLM 回復個性描述，可選。例如：'你是一位專業的數據分析師，要用精準專業的詞彙進行分析。'")
    enable_llm_analysis: Optional[bool] = Field(True, description="是否進行 LLM 分析，預設為 True")
    fH_prompt_alias_level: Optional[int] = Field(3, description="個性的使用等級，預設為2；初步有1~3三種選項，越高越初階")

@app.post("/evaluationPlanTask")
async def evaluationPlanTask(request: EvaluationFitRequest):
    """
    評估模型適配性
    """
    try:
        result = {}
        # 處理 stamps：如果為 None，則使用空列表
        base_stamp = f'Proj_{request.plan_id}-{request.seq_no}'
        stamps = [base_stamp] + (request.stamps if request.stamps else [])
        if not LOGger.isinstance_not_empty(request.output_dir, str):
            output_dir = os.path.join(m_ExportSetPath, str(request.plan_id), str(request.seq_no))
        if not ae.evaluationFit(request.plan_id, request.seq_no, ret=result, stamps=stamps, output_dir=output_dir):
            error_message = result.get('message', '未知錯誤')
            raise HTTPException(status_code=500, detail=f"評估模型適配性失敗: {error_message}")
        
        return {
            "success": True,
            "message": "評估模型適配性完成",
            "result": result
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['evaluationPlanTask'])
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/evaluationPlanTaskLLM")
async def evaluationFitAnalysis(request: EvaluationFitRequest):
    """
    LLM評估模型適配性分析
    """
    try:
        result = {}
        # 處理 stamps：如果為 None，則使用空列表
        base_stamp = f'Proj_{request.plan_id}-{request.seq_no}'
        stamps = [base_stamp] + (request.stamps if request.stamps else [])
        if not LOGger.isinstance_not_empty(request.output_dir, str):
            output_dir = os.path.join(m_ExportSetPath, str(request.plan_id), str(request.seq_no), "Analysis")
        if not ae.evaluationFitAnalysis(request.plan_id, request.seq_no, ret=result, stamps=stamps, output_dir=output_dir, provider=request.provider, model=request.model, user_prompt_template=request.user_prompt_template, system_prompt=request.system_prompt, prompt_temperature=request.prompt_temperature, personality=request.personality, enable_llm_analysis=request.enable_llm_analysis, fH_prompt_alias_level=request.fH_prompt_alias_level, md_output_dir=output_dir):
            error_message = result.get('message', '未知錯誤')
            raise HTTPException(status_code=500, detail=f"LLM評估模型適配性分析失敗: {error_message}")
        
        # 提取 token 和費用信息
        token_summary = result.get('token_summary')
        billing_summary = result.get('billing_summary')
        
        return {
            "success": True,
            "message": "LLM評估模型適配性分析完成",
            "result": result,
            "token_summary": token_summary,
            "billing_summary": billing_summary
        }
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['evaluationPlanTaskLLM'])
        raise HTTPException(status_code=500, detail=str(e))


# 檔案預覽和轉換相關端點
class FilePreviewRequest(BaseModel):
    filePath: str
    sheet: Optional[int] = 0  # xlsx 檔案需要指定工作表
    preview_rows: Optional[int] = 5  # 預覽行數，預設5行

class FileToPklRequest(BaseModel):
    filePath: str
    sheet: Optional[int] = 0  # xlsx 檔案需要指定工作表
    protocol: Optional[int] = 4  # pickle protocol 版本
    output_dir: Optional[str] = "apiOutput"  # 輸出目錄，預設 apiOutput
    output_path: Optional[str] = None  # 輸出檔案路徑（相對於 output_dir），None 表示自動生成

class PklProtocolConvertRequest(BaseModel):
    filePath: str
    protocol: Optional[int] = 4  # 目標 pickle protocol 版本
    output_dir: Optional[str] = "apiOutput"  # 輸出目錄，預設 apiOutput
    output_path: Optional[str] = None  # 輸出檔案路徑（相對於 output_dir），None 表示覆蓋原檔案名

@app.post("/file/preview")
async def preview_file(request: FilePreviewRequest):
    """
    偵測檔案欄位並預覽前N項資料
    
    - 支援 .xlsx, .xls, .csv, .pkl 格式
    - 如果是 xlsx 需要有 sheet 參數
    - 值使用 DFP.parse(x, digit=4) 轉成字串並截取前20字元
    """
    try:
        import pandas as pd
        import pickle
        
        # 解析檔案路徑
        if not os.path.isabs(request.filePath):
            source_dir = config.get('source_dir', '.')
            full_path = os.path.join(source_dir, request.filePath)
            if not os.path.exists(full_path):
                reference_dirs = config.get('referenceDirs', [])
                for ref_dir in reference_dirs:
                    test_path = os.path.join(ref_dir, request.filePath)
                    if os.path.exists(test_path):
                        full_path = test_path
                        break
                else:
                    raise HTTPException(status_code=404, detail=f"檔案不存在: {request.filePath}")
        else:
            full_path = request.filePath
        
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail=f"檔案不存在: {full_path}")
        
        # 檢查檔案類型
        file_ext = os.path.splitext(full_path)[1].lower()
        if file_ext not in ['.xlsx', '.xls', '.csv', '.pkl']:
            raise HTTPException(status_code=400, detail=f"不支援的檔案格式: {file_ext}，僅支援 .xlsx, .xls, .csv, .pkl")
        
        # 載入資料
        if file_ext in ['.xlsx', '.xls', '.pkl']:
            data = DFP.import_data(full_path, sht=request.sheet)
        else:  # .csv
            data = DFP.import_data(full_path)
        
        if data is None or data.empty:
            raise HTTPException(status_code=400, detail="無法載入資料或資料為空")
        
        print(data.head())
        # 預覽資料處理函數
        def format_value(x):
            """格式化值：使用 DFP.parse 轉成字串，digit=4，並截取前20字元"""
            try:
                parsed = DFP.parse(x, digit=4)
                result = str(parsed)[:20]
                return result
            except:
                return str(x)[:20] if x is not None else ''

        # 欄名處理函數：避免 NaN/Inf 欄名導致 JSON 編碼失敗
        def format_column_name(col):
            try:
                if isinstance(col, (float, np.floating)):
                    if math.isnan(col):
                        return "<nan_column>"
                    if math.isinf(col):
                        return "<inf_column>"
                    return str(float(col))
                return str(col)
            except Exception:
                return "<column_error>"
        
        # 預覽前 N 行
        preview_count = min(request.preview_rows, len(data))
        preview_data = data.head(preview_count)
        
        # 格式化預覽資料
        preview_formatted = {}
        for col in preview_data.columns:
            col_name = format_column_name(col)
            preview_formatted[col_name] = [
                format_value(val) for val in preview_data[col].head(preview_count).tolist()
            ]
        
        # 欄位資訊
        columns_info = []
        for col in data.columns:
            col_info = {
                'name': format_column_name(col),
                'dtype': str(data[col].dtype),
                'non_null_count': int(data[col].notna().sum()),
                'null_count': int(data[col].isna().sum()),
                'unique_count': int(data[col].nunique()),
                'sample_values': [format_value(val) for val in data[col].dropna().head(3).tolist()]
            }
            columns_info.append(col_info)

        response_data = {
            "success": True,
            "file_path": full_path,
            "file_type": file_ext,
            "shape": {
                "rows": int(data.shape[0]),
                "columns": int(data.shape[1])
            },
            "columns": columns_info,
            "preview": {
                "rows_shown": preview_count,
                "data": preview_formatted
            }
        }
        return safe_serialize(response_data, max_depth=8)
        
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['preview_file'])
        raise HTTPException(status_code=500, detail=f"預覽檔案失敗: {str(e)}")

@app.post("/file/convert-to-pkl")
async def convert_file_to_pkl(request: FileToPklRequest):
    """
    將 xlsx 或 csv 檔案轉換為 pkl 檔案
    
    - 支援 .xlsx, .xls, .csv 格式
    - xlsx 需要指定 sheet 參數
    - protocol 預設為 4
    """
    try:
        import pickle
        import pandas as pd
        
        # 解析檔案路徑
        if not os.path.isabs(request.filePath):
            source_dir = config.get('source_dir', '.')
            full_path = os.path.join(source_dir, request.filePath)
            if not os.path.exists(full_path):
                reference_dirs = config.get('referenceDirs', [])
                for ref_dir in reference_dirs:
                    test_path = os.path.join(ref_dir, request.filePath)
                    if os.path.exists(test_path):
                        full_path = test_path
                        break
                else:
                    raise HTTPException(status_code=404, detail=f"檔案不存在: {request.filePath}")
        else:
            full_path = request.filePath
        
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail=f"檔案不存在: {full_path}")
        
        # 檢查檔案類型
        file_ext = os.path.splitext(full_path)[1].lower()
        if file_ext not in ['.xlsx', '.xls', '.csv']:
            raise HTTPException(status_code=400, detail=f"不支援的檔案格式: {file_ext}，僅支援 .xlsx, .xls, .csv")
        
        # 載入資料
        if file_ext in ['.xlsx', '.xls']:
            data = DFP.import_data(full_path, sht=request.sheet)
        else:
            data = DFP.import_data(full_path)
        
        if data is None or data.empty:
            raise HTTPException(status_code=400, detail="無法載入資料或資料為空")
        
        # 決定輸出路徑
        output_dir = request.output_dir if request.output_dir else "apiOutput"
        
        if request.output_path:
            # 如果指定了 output_path
            if os.path.isabs(request.output_path):
                # 絕對路徑，直接使用
                output_path = request.output_path
            else:
                # 相對路徑，相對於 output_dir
                output_path = os.path.join(output_dir, request.output_path)
        else:
            # 自動生成：將原檔名改為 .pkl，放在 output_dir 下
            base_name = os.path.splitext(os.path.basename(full_path))[0]
            output_path = os.path.join(output_dir, f"{base_name}.pkl")
        
        # 確保輸出目錄存在
        output_dir_path = os.path.dirname(output_path)
        if output_dir_path and not os.path.exists(output_dir_path):
            os.makedirs(output_dir_path, exist_ok=True)
        
        # 儲存為 pkl 檔案
        with open(output_path, 'wb') as f:
            pickle.dump(data, f, protocol=request.protocol)
        
        return {
            "success": True,
            "message": f"檔案已成功轉換為 pkl 格式",
            "source_file": full_path,
            "output_file": output_path,
            "protocol": request.protocol,
            "shape": {
                "rows": int(data.shape[0]),
                "columns": int(data.shape[1])
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['convert_file_to_pkl'])
        raise HTTPException(status_code=500, detail=f"轉換檔案失敗: {str(e)}")

@app.post("/file/convert-pkl-protocol")
async def convert_pkl_protocol(request: PklProtocolConvertRequest):
    """
    轉換 pkl 檔案的 protocol 版本
    
    - 僅支援 .pkl 檔案
    - protocol 預設為 4
    - 其他格式檔案會被拒絕
    """
    try:
        import pickle
        
        # 解析檔案路徑
        if not os.path.isabs(request.filePath):
            source_dir = config.get('source_dir', '.')
            full_path = os.path.join(source_dir, request.filePath)
            if not os.path.exists(full_path):
                reference_dirs = config.get('referenceDirs', [])
                for ref_dir in reference_dirs:
                    test_path = os.path.join(ref_dir, request.filePath)
                    if os.path.exists(test_path):
                        full_path = test_path
                        break
                else:
                    raise HTTPException(status_code=404, detail=f"檔案不存在: {request.filePath}")
        else:
            full_path = request.filePath
        
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail=f"檔案不存在: {full_path}")
        
        # 檢查檔案類型（只接受 .pkl）
        file_ext = os.path.splitext(full_path)[1].lower()
        if file_ext != '.pkl':
            raise HTTPException(status_code=400, detail=f"僅支援 .pkl 檔案，收到: {file_ext}")
        
        # 讀取 pkl 檔案
        try:
            with open(full_path, 'rb') as f:
                data = pickle.load(f)
        except Exception as e:
            LOGger.exception_process(e, logfile='', stamps=['convert_pkl_protocol'])
            raise HTTPException(status_code=400, detail=f"無法讀取 pkl 檔案: {str(e)}")
        
        # 決定輸出路徑
        output_dir = request.output_dir if request.output_dir else "apiOutput"
        
        if request.output_path:
            # 如果指定了 output_path
            if os.path.isabs(request.output_path):
                # 絕對路徑，直接使用
                output_path = request.output_path
            else:
                # 相對路徑，相對於 output_dir
                output_path = os.path.join(output_dir, request.output_path)
        else:
            # 自動生成：保持原檔名，放在 output_dir 下
            base_name = os.path.basename(full_path)
            output_path = os.path.join(output_dir, base_name)
        
        # 確保輸出目錄存在
        output_dir_path = os.path.dirname(output_path)
        if output_dir_path and not os.path.exists(output_dir_path):
            os.makedirs(output_dir_path, exist_ok=True)
        
        # 以新 protocol 儲存
        with open(output_path, 'wb') as f:
            pickle.dump(data, f, protocol=request.protocol)
        
        # 獲取資料形狀（如果是 DataFrame）
        shape_info = None
        if hasattr(data, 'shape'):
            shape_info = {
                "rows": int(data.shape[0]),
                "columns": int(data.shape[1]) if len(data.shape) > 1 else int(data.shape[0])
            }
        
        return {
            "success": True,
            "message": f"pkl 檔案 protocol 已轉換為 {request.protocol}",
            "source_file": full_path,
            "output_file": output_path,
            "protocol": request.protocol,
            "shape": shape_info
        }
        
    except HTTPException:
        raise
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=['convert_pkl_protocol'])
        raise HTTPException(status_code=500, detail=f"轉換 pkl protocol 失敗: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    
    # 添加離散數據聚合端點
    try:
        from src.api_integration import add_discrete_aggregation_endpoints
        add_discrete_aggregation_endpoints(app)
        print("✓ 離散數據聚合端點已添加")
    except ImportError as e:
        print(f"⚠ 離散數據聚合功能載入失敗: {e}")
    
    # 從配置檔案取得主機和埠號設定
    host = config.get("Host_IP", "127.0.0.1")
    port = config.get("Host_Port", 8000)
    
    print(f"正在啟動 DataAnalysis API 服務...")
    print(f"服務地址: http://{host}:{port}")
    print(f"API 文檔: http://{host}:{port}/docs")
    print(f"配置資訊: http://{host}:{port}/config")
    print(f"來源目錄: {config.get('source_dir', '.')}")
    print(f"參考目錄: {config.get('referenceDirs', [])}")
    
    uvicorn.run(app, host=host, port=port)
